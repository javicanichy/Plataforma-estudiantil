import os
import re
import io

from dotenv import load_dotenv
from datetime import datetime, date, timedelta
from functools import wraps
from werkzeug.middleware.proxy_fix import ProxyFix
from email_validator import validate_email, EmailNotValidError
from io import BytesIO
from docx import Document
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from fpdf import FPDF
from dateutil import parser

from flask import (
    Flask, jsonify, request, redirect, url_for, send_file, send_from_directory,
    render_template, send_file, session, abort, flash, Blueprint, current_app, make_response
)

import unicodedata
import json
import logging
import smtplib
import bleach
import uuid
import email_validator
import secrets
import string
import random
import pandas as pd
import psutil
import shutil

from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_cors import CORS
from flask_mail import Message, Mail, Connection
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from sqlalchemy import or_, inspect, text
from sqlalchemy.orm import joinedload
from sqlalchemy.orm.attributes import flag_modified
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, BooleanField, FileField, SubmitField


from forms import (
    NoticiaForm, PerfilForm, CambiarContrasenaForm, FileAllowed_PERFILES_EXIT, Email, EqualTo, DataRequired, 
    Length, Optional, ValidationError, PasswordField, DebateForm, BibliotecaForm, LibroFisicoForm,
    SolicitudPrestamoForm, BuzonAyudaForm, OpinionForm, SelectividadForm, SelectividadForm, MatriculaForm,
    )

from models import (
    SecretariaActa, db, Usuario, Estudiante, Nota, Mensaje, Evento, Evento, Debate, Notificacion, Administrador, Comentario,
    CodigoEstudiante, Asignatura, Noticia, Debate, Notificacion, Comentario,
    Biblioteca, Buzon, OpinionSelectividad, Selectividad, SolicitudMatricula, Expediente, Profesor,
    Directivo, SecretariaActa, Carpeta, DocumentoArchivo, Secretaria, AnuncioDirectivo, DocumentoRecibido
    )
    
from config import Config
from flask_migrate import Migrate



"""=========================================
 ESTRUCTURA BÁSICA:
 1. Base de datos → PostgreSQL
 2. Backend → Flask + SQLAlchemy + CORS
 3. Frontend → HTML + CSS + JS (Fetch API)
 ========================================="""



# ==========================================================
# CONFIGURACIÓN GENERAL
# ==========================================================

# Cargar variables de entorno
load_dotenv()

# Directorios
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

# Configurar app de Flask
app = Flask(__name__, static_folder=STATIC_DIR, template_folder=TEMPLATES_DIR)
# APLICAR PROXY FIX: Corrige la gestión de sesiones bajo el proxy de Render (HTTPS)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# Configuarar db
# db = SQLAlchemy()

# Configuracion desde config.py
app.config.from_object(Config)

# SEGURIDAD Y SESIONES
#-----------------------------------------
app.secret_key = app.config['SECRET_KEY']

# Permite hasta 32 Megabytes de subida (suficiente para todos los PDFs y fotos)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

# Indica a Flask que confíe en los proxies (necesario en Render/servidores en la nube)
app.config['PREFERRED_URL_SCHEME'] = 'https'
#Forzar que la cookie de sesión solo se envíe sobre HTTPS, esto resuelve el problema de la sesión en el navegador de Render
app.config['SESSION_COOKIE_SECURE'] = True

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
#-----------------------------------------

#------ CONFIGURACION DE CORREOS--------#
# MANEJAR CORREOS ELECTRONICOS. BUZON DE AYUDA
app.config['MAIL_SERVER'] = 'smtp.gmail.com'        # servidor SMTP, puedes cambiarlo según tu correo
app.config['MAIL_PORT'] = 587                       # puerto TLS
app.config['MAIL_USE_TLS'] = True                   # usar TLS
app.config['MAIL_USE_SSL'] = False                  # no usar SSL porque usamos TLS
app.config['MAIL_USERNAME'] = 'tucorreo@gmail.com'  # tu correo
app.config['MAIL_PASSWORD'] = 'tu_contraseña_app'   # contraseña de aplicación o normal (Gmail necesita app password)
app.config['MAIL_DEFAULT_SENDER'] = 'tucorreo@gmail.com'  # remitente por defecto

# Inicializar Flask-Mail
mail = Mail(app)



# SEGUNDA CONFIGURACION DEL CORREO. SECRETARIA-MATRICULA, ESTUDIANTES
# Definimos estas variables por separado para no sobrescribir app.config
"""CORREO_MATRICULAS_USER = 'secretaria-matriculas@gmail.com' 
CORREO_MATRICULAS_PASS = 'tu_segunda_contraseña_app'
CORREO_MATRICULAS_SERVER = 'smtp.gmail.com'
CORREO_MATRICULAS_PORT = 587"""


# TERCERA CONFIGURACION DEL CORREO. REGISTRO DE PROFESORES
# Definimos estas variables por separado para no sobrescribir app.config
"""CORREO_MATRICULAS_USER = 'secretaria-matriculas@gmail.com' 
CORREO_MATRICULAS_PASS = 'tu_segunda_contraseña_app'
CORREO_MATRICULAS_SERVER = 'smtp.gmail.com'
CORREO_MATRICULAS_PORT = 587"""



# CUARTA CONFIGURACION DEL CORREO. SUSPENCION DE CUENTAS
# Definimos estas variables por separado para no sobrescribir app.config
"""CORREO_MATRICULAS_USER = 'secretaria-matriculas@gmail.com' 
CORREO_MATRICULAS_PASS = 'tu_segunda_contraseña_app'
CORREO_MATRICULAS_SERVER = 'smtp.gmail.com'
CORREO_MATRICULAS_PORT = 587"""



# QUITA CONFIGURACION DEL CORREO. RECUPERACION DE CUENTAS
# Definimos estas variables por separado para no sobrescribir app.config
"""CORREO_MATRICULAS_USER = 'secretaria-matriculas@gmail.com' 
CORREO_MATRICULAS_PASS = 'tu_segunda_contraseña_app'
CORREO_MATRICULAS_SERVER = 'smtp.gmail.com'
CORREO_MATRICULAS_PORT = 587"""


# CONFIGURACIÓN DE CORREO (TEMPORAL PARA MAILHOG).
# ==========================================
CORREO_MATRICULAS_SERVER = 'localhost'
CORREO_MATRICULAS_PORT = 1025
CORREO_MATRICULAS_USER = 'test@unge.gq'
CORREO_MATRICULAS_PASS = '' 
# ==========================================

# Arancar la abse de datos
db.init_app(app)
CORS(app)

# Mirgraciones
migrate = Migrate(app, db)





# Definición del Blueprint
eventos_bp = Blueprint('eventos', __name__, url_prefix='/api/eventos')


# ===== INYECTAR VARIABLES GLOBALES EN TODOS LOS TEMPLATES (ANTES DE CUALQUIER RUTA) =====
@app.context_processor
def inyectar_contexto():
    """Inyecta datos del usuario en TODOS los templates automáticamente"""
    return {
        'logueado': current_user.is_authenticated,
        'usuario': current_user.nombre if current_user.is_authenticated else 'Invitado',
        'rol': current_user.rol if current_user.is_authenticated else None,
        'usuario_id': current_user.id if current_user.is_authenticated else None,
        'current_user': current_user  # Esto te permite usar {{ current_user.apellidos }} en cualquier HTML
    }

# ==========================================================
# MANEJO DE ARCHIVOS
# ==========================================================

# Carpeta donde se guardarán los archivos subidos
UPLOAD_FOLDER = app.config['UPLOAD_FOLDER']
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_IMAGE_EXT = {'jpg', 'jpeg', 'png', 'gif'}
ALLOWED_VIDEO_EXT = {'mp4', 'mov', 'avi'}
ALLOWED_EXT = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT



# Carpeta base para todos los archivos subidos
UPLOADS_DIR = os.path.join(STATIC_DIR, 'uploads')
os.makedirs(UPLOADS_DIR, exist_ok=True)  # Crear uploads si no existe

def guardar_archivo(archivo, categoria):
    """
    Guarda un archivo subido en la carpeta correcta según categoría.
    Devuelve la ruta relativa para guardar en DB.
    """
    if not archivo:
        return None

    # Crear carpeta de categoría si no existe
    carpeta_destino = os.path.join(UPLOADS_DIR, categoria)
    os.makedirs(carpeta_destino, exist_ok=True)

    # Nombre seguro
    filename = secure_filename(archivo.filename)

    # Ruta final en el sistema
    ruta_guardado = os.path.join(carpeta_destino, filename)
    archivo.save(ruta_guardado)

    # Ruta relativa para templates / URLs
    return f'uploads/{categoria}/{filename}'



# ==========================================================
# DECORADORES PROFESIONALES. Rol restrintion
# ==========================================================


# 1. Decorador para verificar si está logueado
def requiere_login(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # En lugar de buscar en 'session', preguntamos a Flask-Login
        if not current_user.is_authenticated:
            flash("Por favor, inicia sesión para acceder.", "warning")
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

# 2. Decorador para verificar el ROL
from functools import wraps
from flask import flash, redirect, url_for
from flask_login import current_user

def requiere_rol(roles_permitidos):
    # Si nos pasan un solo string (un solo rol), lo convertimos en lista para que la lógica funcione igual
    if isinstance(roles_permitidos, str):
        roles_permitidos = [roles_permitidos]
        
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # 1. Verificar si está autenticado
            if not current_user.is_authenticated:
                flash("Por favor, inicia sesión para acceder.", "warning")
                return redirect(url_for('login'))
            
            # 2. Verificar si el rol del usuario está DENTRO de la lista permitida
            if current_user.rol not in roles_permitidos:
                roles_texto = ", ".join(roles_permitidos)
                flash(f"Acceso denegado. Se requiere uno de estos roles: {roles_texto}.", "danger")
                return redirect(url_for('inicio'))
                
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ===========================================================================
# CONFIGURACIÓN DE CARRERAS Y CURSOS (Diccionario)
CARRERAS_INFO = {
    'Grado Medicina': {'años': 6},
    'Grado II Enfermería': {'años': 4},
    'Grado Fisioterapia': {'años': 4},
    'Grado I Enfermería': {'años': 3},
    'Grado I Ginecobstetrica': {'años': 3},
    'Grado I Laboratorio': {'años': 3},
    'Grado I Epidemiologia': {'años': 3},
    'Grado I Imagenologia': {'años': 3},
    'Grado I Anestesiologia': {'años': 3}
}


# ==========================================================
# RUTAS PÚBLICAS
# ==========================================================
@app.route('/', methods=['GET', 'POST'])
def inicio():
    form = BuzonAyudaForm()
    
    if form.validate_on_submit():
        tipo = form.tipo_consulta.data  # 'matricula' o 'general'
        ruta_archivo = None
        
        # 1. Guardar el archivo PDF si existe
        if form.archivo.data:
            ruta_completa = guardar_archivo(form.archivo.data, 'matriculas_docs')
            # Extraemos solo el nombre: "Seleccion_Masculina_FCS.pdf"
            nombre_solo = os.path.basename(ruta_completa)
            ruta_archivo = nombre_solo

        try:
            # 2. PROCESO EXCLUSIVO PARA MATRÍCULA/ADMISIÓN
            if tipo == 'matricula':
                dip_ingresado = form.dip.data
                # Buscamos al estudiante en la tabla de solicitudes
                estudiante = SolicitudMatricula.query.filter_by(dni_numero=dip_ingresado).first()
                
                if estudiante:
                    # Vinculamos el documento directamente a su ficha de matrícula
                    if ruta_archivo:
                        estudiante.doc_ficha_matricula = ruta_archivo
                        estudiante.estado = "Admitido" # Marcamos para revisión del admin
                    
                    # GUARDAMOS EL MENSAJE EN LA NUEVA COLUMNA
                    estudiante.mensaje_buzon = form.mensaje.data

                    # Creamos el registro en el buzón con referencia al alumno
                    nueva_consulta = Buzon(
                        tipo_consulta='matricula',
                        nombre=f"{estudiante.nombre} {estudiante.apellidos}",
                        dip=dip_ingresado,
                        correo=form.correo.data,
                        mensaje=f"[TRÁMITE ADMISIÓN] {form.mensaje.data}",
                        archivo=ruta_archivo
                    )
                    flash(f"Documento vinculado al DIP {dip_ingresado} correctamente.", "success")
                else:
                    # Si el DIP no existe, el admin lo recibirá como alerta de DIP no encontrado
                    nueva_consulta = Buzon(
                        tipo_consulta='matricula',
                        nombre="DIP NO REGISTRADO",
                        dip=dip_ingresado,
                        correo=form.correo.data,
                        mensaje=f"ALERTA: Intento de envío con DIP inexistente: {form.mensaje.data}",
                        archivo=ruta_archivo
                    )
                    flash("El DIP no coincide, pero su mensaje fue enviado al administrador.", "warning")

            # 3. PROCESO PARA CONSULTA GENERAL
            else:
                nueva_consulta = Buzon(
                    tipo_consulta='general',
                    nombre=form.nombre.data,
                    correo=form.correo.data,
                    mensaje=form.mensaje.data,
                    archivo=ruta_archivo
                )
                flash("Consulta general enviada con éxito.", "success")

            db.session.add(nueva_consulta)
            db.session.commit()
            return redirect(url_for('inicio', _anchor='buzon'))

        except Exception as e:
            db.session.rollback()
            print(f"Error en el proceso: {e}")
            flash("Error al procesar la solicitud.", "danger")
            
    return render_template('index.html', form=form, noticias=Noticia.query.all())

@app.route('/ver-documento/<filename>')
def ver_documento(filename):
    folder = os.path.join(app.root_path, 'static', 'uploads', 'matriculas_docs')
    
    # Comprobamos si el archivo existe físicamente
    if not os.path.exists(os.path.join(folder, filename)):
        return "Archivo no encontrado", 404

    response = make_response(send_from_directory(folder, filename))
    
    # Estos encabezados son los que convencen a Chrome de mostrarlo
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline'
    # Esta línea elimina restricciones de seguridad para el visor
    response.headers['X-Frame-Options'] = 'ALLOWALL' 
    
    return response

@app.route('/admin/buzon-general')
# @login_required  <-- Descomenta esto si usas LoginManager
def buzon_general():
    # Filtramos solo las que son de tipo 'general'
    consultas = Buzon.query.filter_by(tipo_consulta='general').order_by(Buzon.fecha.desc()).all()
    return render_template('buzon.html', consultas=consultas)

@app.route('/login')
def login_page():
    return render_template('login.html', body_class="fondo-login")

@app.route('/registro', methods=['GET', 'POST'])
def registro_page():
    return render_template('registro.html')

@app.route("/noticias")
def noticias_page():
    noticias = Noticia.query.order_by(Noticia.fecha.desc()).all()
    return render_template("lista_noticia.html", noticias=noticias)

@app.route('/contacto')
def contacto_page():
    return render_template('contacto.html')


@app.route('/requisitos')
def requisitos():
    return render_template('requisitos.html')

@app.route('/asignaturas')
def asignaturas_page():
    return render_template('asignaturas.html')

@app.route('/calendario')
def calendario_page():
    return render_template('calendario.html', body_class="calendario-page")

@app.route('/perfil')
@login_required
def perfil_page():
    return redirect(url_for('ver_perfil', usuario_id=current_user.id))



# ==========================================================
# Inject current_user into templates
@app.context_processor
def inject_user():
    if 'usuario_id' in session:
        usuario = Usuario.query.get(session['usuario_id'])
        return {'current_user': usuario}
    return {'current_user': None}

@app.route("/biblioteca")
def biblioteca_page():
    # Libros de TFG visibles para todos
    filtro_titulacion = request.args.get("titulacion", None)

    query_tfg = Biblioteca.query.filter_by(tipo_libro="tfg", publico=True)

    if filtro_titulacion and filtro_titulacion != "":
        query_tfg = query_tfg.filter_by(titulacion=filtro_titulacion)

    tfg_publicos = query_tfg.order_by(Biblioteca.titulacion.asc(), Biblioteca.fecha_creacion.desc()).all()

    return render_template(
        "biblioteca.html",
        tfg_publicos=tfg_publicos,
        filtro_titulacion=filtro_titulacion,
        body_class="biblioteca-page"
    )

# ==========================================================
# API: SESIÓN DE USUARIO
# ==========================================================

@app.route('/api/login', methods=['POST'])
@app.route('/login', methods=['POST'])
def api_login():
    if not request.is_json:
        return jsonify({'mensaje': 'Se requiere JSON'}), 400

    data = request.get_json()
    correo_inst = (data.get('correo') or data.get('email') or '').strip().lower()
    clave = data.get('clave') or data.get('password') or ''

    usuario = Usuario.query.filter_by(correo_institucional=correo_inst).first()

    if not usuario:
        return jsonify({
            'ok': False, 
            'mensaje': 'Correo institucional no registrado o incorrecto'
        }), 404

    # 1. VERIFICACIÓN DE CUENTA ACTIVA (Control Disciplinario)
    if not getattr(usuario, 'activo', True): # Si 'activo' es False, bloqueamos
        return jsonify({
            'ok': False,
            'mensaje': 'Su cuenta ha sido desactivada por control disciplinario. Contacte con administración.'
        }), 403

    # 2. VERIFICACIÓN DE EXPIRACIÓN TEMPORAL
    if usuario.fecha_expiracion and datetime.utcnow() > usuario.fecha_expiracion:
        # Si ya expiró, la desactivamos automáticamente en la BD
        usuario.activo = False
        db.session.commit()
        return jsonify({
            'ok': False,
            'mensaje': 'Su periodo de acceso ha expirado. Debe renovar su cuenta con el directivo.'
        }), 403

    # --- NUEVA VALIDACIÓN DE ROL ---
    if usuario.rol in ['admin', 'administrador']:
        return jsonify({
            'ok': False, 
            'mensaje': 'Este acceso es exclusivo para alumnos. Use el portal administrativo.'
        }), 403 
    # -------------------------------

    if not check_password_hash(usuario.password_hash, clave):
        return jsonify({'ok': False, 'mensaje': 'Contraseña incorrecta'}), 401

    login_user(usuario, remember=True) 

    # Guardar en sesión
    session['usuario_id'] = usuario.id
    session['nombre'] = usuario.nombre
    session['rol'] = usuario.rol

    return jsonify({
        'ok': True,
        'mensaje': 'Acceso concedido',
        'usuario': {
            'id': usuario.id, 
            'nombre': usuario.nombre, 
            'correo': usuario.correo_institucional,
            'talento': usuario.talento
        }
    }), 200


@app.route('/api/usuario_sesion', methods=['GET'])
def usuario_sesion():
    # Ahora usamos current_user de Flask-Login (mucho más seguro)
    
    if not current_user.is_authenticated:
        return jsonify({'logueado': False})
        
    return jsonify({
        'logueado': True,
        'id': current_user.id,
        'nombre': current_user.nombre,
        'rol': current_user.rol,
        'talento': getattr(current_user, 'talento', None) # Ya incluimos el talento
    })



# ===========================================================
# RUTAS DEL CALENDARIO (API)
# ===========================================================

# 1. Obtener eventos (Para que aparezcan al cargar el calendario)
@app.route('/api/eventos', methods=['GET'])
@login_required
def get_eventos():
    try:
        eventos = Evento.query.all()
        # Usamos tu método to_dict() que ya mapea titulo -> title correctamente
        return jsonify([e.to_dict() for e in eventos])
    except Exception as e:
        print(f"Error al leer eventos: {e}")
        return jsonify([]), 500
        

# 2. Crear evento (Solo Admin y Directivo)

@app.route('/api/eventos', methods=['POST'])
@login_required
def add_evento():
    if current_user.rol not in ['admin', 'directivo']:
        return jsonify({'error': 'No autorizado'}), 403

    data = request.get_json()
    try:
        # 1. Limpiamos las fechas que vienen de FullCalendar (JS)
        # Reemplazamos 'Z' por nada para que Python lo lea como DateTime
        start_dt = datetime.fromisoformat(data['start'].replace('Z', ''))
        end_dt = None
        if data.get('end'):
            end_dt = datetime.fromisoformat(data['end'].replace('Z', ''))

        # 2. Creamos el objeto usando TUS nombres de columna
        nuevo_evento = Evento(
            titulo=data['title'],      # Tu columna es 'titulo'
            start=start_dt,            # Tu columna es 'start'
            descripcion=data.get('description'),
            end=end_dt,                # Tu columna es 'end'
            all_day=data.get('allDay', True), # Tu columna es 'all_day'
            tipo=data.get('tipo', 'general'), # Tu columna es 'tipo'
            usuario_id=current_user.id
        )

        db.session.add(nuevo_evento)
        db.session.commit()
        
        # 3. Usamos tu función to_dict() para responderle al calendario
        return jsonify(nuevo_evento.to_dict()), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Error al guardar evento: {e}") # Para que lo veas en consola
        return jsonify({'error': str(e)}), 500

# 3. Eliminar evento (Solo Admin y Directivo)
@app.route('/api/eventos/<int:id>', methods=['DELETE'])
@login_required
def delete_evento(id):
    # Seguridad: solo admin y directivo
    if current_user.rol not in ['admin', 'directivo']:
        return jsonify({'error': 'No autorizado'}), 403
        
    evento = Evento.query.get_or_404(id)
    try:
        db.session.delete(evento)
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



# ==========================================================
# API: CREACION DE CORREOS CORPORATIVOS PARA ESTUDIANTES
# ==========================================================
def crear_correo_unge_estandar(nombre, apellidos):
    """
    Genera correo: inicialesNombre.primerApellido.año.cienciasdelasalud@unge.gq
    Ejemplo: Isaias Nkogo + Esono Aviri -> in.esono.2025.cienciasdelasalud@unge.gq
    """
    def normalizar(texto):
        if not texto: return ""
        # Elimina acentos y convierte a minúsculas
        texto_norm = unicodedata.normalize('NFD', texto)
        return "".join(c for c in texto_norm if unicodedata.category(c) != 'Mn').lower().strip()

    # 1. Obtener iniciales de TODOS los nombres (Isaias Nkogo -> in)
    nombres_lista = nombre.split()
    iniciales = "".join([normalizar(n)[0] for n in nombres_lista if n])

    # 2. Obtener solo el primer apellido (Esono Aviri -> esono)
    primer_apellido = normalizar(apellidos.split()[0])

    # 3. Año actual
    anio = str(datetime.now().year)

    # 4. Construir base con el estándar fijo de la facultad
    # Resultado: in.esono.2025.cienciasdelasalud
    correo_base = f"{iniciales}.{primer_apellido}.{anio}.cienciasdelasalud"
    dominio = "@unge.gq"
    
    email_final = f"{correo_base}{dominio}"
    
    # 5. Control de duplicados (en caso de que dos alumnos tengan mismas iniciales y apellido)
    contador = 1
    while Usuario.query.filter_by(correo_institucional=email_final).first():
        email_final = f"{correo_base}{contador}{dominio}"
        contador += 1
        
    return email_final



def enviar_codigo_activacion(solicitud, codigo_nuevo, dominio):
    """
    Envía el correo de aprobación con el código de un solo uso,
    manteniendo el diseño institucional de la UNGE.
    """
    email_destino = solicitud.email
    nombre_alumno = f"{solicitud.nombre} {solicitud.apellidos}"
    carrera_alumno = solicitud.carrera

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Secretaría Académica UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = "¡Solicitud Admitida! - Código de Activación de Cuenta"

    # Enlace directo al wizard de registro
    url_registro = f"{dominio}/registro-estudiante"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Facultad de Ciencias de la Salud</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #1b5e20;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 300; letter-spacing: 1px;">SOLICITUD ADMITIDA</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">¡Felicidades, <strong>{nombre_alumno}</strong>!</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Nos complace informarle que su solicitud para la carrera de <strong>{carrera_alumno}</strong> ha sido <strong>APROBADA</strong>. 
                        Ya puede proceder a activar su cuenta de estudiante y generar su correo institucional.
                    </p>

                    <div style="background-color: #f1f8e9; border-radius: 10px; padding: 25px; border: 2px dashed #1b5e20; text-align: center; margin-bottom: 30px;">
                        <p style="margin: 0 0 10px 0; font-size: 13px; color: #1b5e20; font-weight: bold; text-transform: uppercase;">Su código de activación es:</p>
                        <span style="font-size: 32px; font-weight: bold; color: #1b5e20; letter-spacing: 4px;">{codigo_nuevo}</span>
                    </div>

                    <p style="font-size: 15px; color: #455a64; margin-bottom: 25px; text-align: center;">
                        Para activar su cuenta, haga clic en el siguiente botón e introduzca su DIP junto con este código:
                    </p>

                    <div style="text-align: center; margin-bottom: 30px;">
                        <a href="{url_registro}" style="background-color: #1a237e; color: #ffffff; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 16px; display: inline-block;">ACTIVAR MI CUENTA</a>
                    </div>

                    <p style="font-size: 13px; color: #78909c; text-align: center; font-style: italic;">
                        Recuerde que este código es de un solo uso y es personal e intransferible.
                    </p>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Secretaría Académica - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar código de activación: {e}")
        return False


# NO ENVIAR CODIGO DE ESTUDIANTE SI COMPROBANTE DE MATRICULA NO ESTÁ APROBADO
def enviar_rechazo_solicitud(solicitud, motivo, dominio):
    """
    Envía el correo de rechazo o solicitud de corrección,
    manteniendo la línea gráfica de la UNGE.
    """
    email_destino = solicitud.email
    nombre_alumno = f"{solicitud.nombre} {solicitud.apellidos}"
    carrera_alumno = solicitud.carrera

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Secretaría Académica UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = "Acción Requerida: Revisión de su Solicitud de Matrícula"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Facultad de Ciencias de la Salud</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #b71c1c;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 300; letter-spacing: 1px;">SOLICITUD RECHAZADA / PENDIENTE</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a <strong>{nombre_alumno}</strong>,</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Lamentamos informarle que su solicitud para la carrera de <strong>{carrera_alumno}</strong> no ha podido ser procesada debido a inconsistencias en la documentación presentada.
                    </p>

                    <div style="background-color: #fff8e1; border-radius: 10px; padding: 25px; border-left: 5px solid #ff6f00; margin-bottom: 30px;">
                        <p style="margin: 0 0 10px 0; font-size: 13px; color: #e65100; font-weight: bold; text-transform: uppercase;">Observaciones de Secretaría:</p>
                        <p style="font-size: 16px; color: #333; font-style: italic; margin: 0;">"{motivo}"</p>
                    </div>

                    <p style="font-size: 15px; color: #455a64; margin-bottom: 25px;">
                        Para continuar con su proceso de inscripción, es necesario que subsane los errores mencionados. Por favor, póngase en contacto con la administración o realice una nueva solicitud con los documentos correctos.
                    </p>

                    <p style="font-size: 13px; color: #78909c; text-align: center; font-style: italic;">
                        Su expediente permanecerá en estado de pausa hasta que se reciba la corrección requerida.
                    </p>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Secretaría Académica - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar correo de rechazo: {e}")
        return False


# ==========================================================
# API: REGISTRO PARA ESTUDIANTES
# ==========================================================
@app.route('/registro-estudiante', methods=['GET', 'POST'])
def registro_estudiante():
    if current_user.is_authenticated:
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        datos = request.form
        dip_ingresado = datos.get('dip')
        codigo_ingresado = datos.get('codigo_estudiante')

        # 1. Validar Código de Estudiante
        codigo_db = CodigoEstudiante.query.filter_by(
            codigo=codigo_ingresado, 
            estudiante_dip=dip_ingresado,
            usado=False
        ).first()

        if not codigo_db:
            return render_template('registro_fallido.html', mensaje="El código o el DIP no son válidos o ya han sido usados.")

        # 2. Buscar la Solicitud de Matrícula
        solicitud = SolicitudMatricula.query.filter_by(dni_numero=dip_ingresado).first()

        if not solicitud:
            return render_template('registro_fallido.html', mensaje="No se encontró una solicitud de matrícula aprobada para este DIP.")

        # 3. Lógica de creación de correo institucional
        email_inst = crear_correo_unge_estandar(solicitud.nombre, solicitud.apellidos)

        try:
            # 4. Crear el Usuario con TRASPASO AUTOMÁTICO
            nuevo_usuario = Usuario(
                nombre=solicitud.nombre,
                apellidos=solicitud.apellidos,
                sexo=solicitud.sexo, 
                natural_de=solicitud.natural_de,
                distrito_provincia=solicitud.distrito_provincia,
                correo=solicitud.email, 
                correo_institucional=email_inst,
                dip=solicitud.dni_numero,
                telefono=solicitud.telefono,
                fecha_nacimiento=str(solicitud.fecha_nacimiento),
                residencia=solicitud.residencia,
                pais=solicitud.nacionalidad,
                carrera=codigo_db.titulacion_autorizada,
                # --- NUEVOS CAMPOS ---
                talento=datos.get('talento'), # Recogemos el valor del select con iconos
                biografia=datos.get('biografia'),
                rol='estudiante'
            )
            # El método set_password se encarga de que el correo institucional 
            # reconozca la clave mediante el hash de seguridad.
            nuevo_usuario.set_password(datos.get('password'))

            # Marcar código como usado
            codigo_db.usado = True
            
            db.session.add(nuevo_usuario)
            db.session.commit()

            # Inicio de sesión automático tras activar
            login_user(nuevo_usuario)
            
            # Pasamos el usuario al template para mostrar su nuevo correo institucional
            return render_template('registro_exitoso.html', usuario=nuevo_usuario)

        except Exception as e:
            db.session.rollback()
            print(f"Error: {e}") # Para depuración en consola
            return render_template('registro_fallido.html', mensaje=f"Error al procesar el registro: {str(e)}")

    return render_template('registro_wizard.html')


# Mostrar el nombre y apellido asociado a un DIP
@app.route('/verificar-dip/<dip>')
def verificar_dip(dip):
    solicitud = SolicitudMatricula.query.filter_by(dni_numero=dip).first()
    if solicitud:
        return jsonify({
            'encontrado': True,
            'nombre': solicitud.nombre,
            'apellido': solicitud.apellidos
        })
    return jsonify({'encontrado': False})

# PROCESAR SOLICITUD DE CODIGO DE MATRICULA, LO QUE VE EL ADMIN
@app.route('/admin/panel-solicitudes')
@login_required
def panel_admin_solicitudes():
    # 1. Traemos los códigos para saber quién ya fue procesado
    codigos = {c.estudiante_dip: c for c in CodigoEstudiante.query.all()}
    
    # 2. Filtramos: Solo 'Admitido' (esto excluye automáticamente a los 'Rechazada')
    # Y que tengan un mensaje o un documento adjunto
    solicitudes = SolicitudMatricula.query.filter_by(estado='Admitido').all()
    
    solicitudes_visibles = [
        s for s in solicitudes 
        if s.dni_numero not in codigos  # Si ya tiene código, no lo mostramos
    ]

    return render_template('admin_solicitudes.html', 
                           solicitudes=solicitudes_visibles, 
                           codigos=codigos)


# PERMITIR AUTOGENERACION DE CODIGO ESTUDIANTE POR PARTE DE ADMINISTRADOR
# Reutilizamos tu instancia de mail configurada con Mailhog
def generar_codigo_seguro():
    caracteres = string.ascii_uppercase + string.digits
    return ''.join(secrets.choice(caracteres) for _ in range(8))

@app.route('/admin/aprobar-solicitud/<int:id>')
def aprobar_solicitud(id):
    solicitud = SolicitudMatricula.query.get_or_404(id)
    
    # 1. Generar código
    codigo_txt = f"UNGE-{generar_codigo_seguro()}"
    dominio_actual = request.host_url.rstrip('/') 

    nuevo_permiso = CodigoEstudiante(
        codigo=codigo_txt,
        estudiante_dip=solicitud.dni_numero,
        titulacion_autorizada=solicitud.carrera,
        usado=False
    )
    
    try:
        db.session.add(nuevo_permiso)
        
        # 2. ENVIAR EL NUEVO CORREO ESTRUCTURADO
        envio_exitoso = enviar_codigo_activacion(solicitud, codigo_txt, dominio_actual)
        
        if envio_exitoso:
            db.session.commit()
            return jsonify({"status": "success", "codigo": codigo_txt})
        else:
            return jsonify({"status": "error", "message": "Error al conectar con Mailhog"})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)})

# NO ENVIAR CODIGO DE ESTUDIANTE SI COMPROBANTE DE MATRICULA NO ESTÁ APROBADO
@app.route('/admin/rechazar-solicitud/<int:id>', methods=['POST'])
@login_required
def rechazar_solicitud(id):
    solicitud = SolicitudMatricula.query.get_or_404(id)
    datos = request.get_json()
    motivo_txt = datos.get('motivo', 'Documento incorrecto.')
    
    try:
        # 1. Enviamos el correo de notificación
        envio_ok = enviar_rechazo_solicitud(solicitud, motivo_txt, request.host_url.rstrip('/'))
        
        if envio_ok:
            # 2. CAMBIO DE ESTADO: Fundamental para que desaparezca al recargar (F5)
            # Al dejar de ser "Admitido", el filtro del panel ya no lo incluirá.
            solicitud.estado = "Rechazada"

            # 3. LIMPIEZA DE DATOS:
            # Borramos el mensaje de la columna que acabas de crear en SolicitudMatricula
            solicitud.mensaje_buzon = None
            
            # Borramos también el registro físico de la tabla Buzon (histórico)
            mensaje_a_borrar = Buzon.query.filter_by(dip=solicitud.dni_numero).first()
            if mensaje_a_borrar:
                db.session.delete(mensaje_a_borrar)
            
            db.session.commit()
            return jsonify({"status": "success", "message": "Estudiante rechazado y retirado del panel"})
        else:
            return jsonify({"status": "error", "message": "No se pudo enviar el correo de rechazo"}), 500
            
    except Exception as e:
        db.session.rollback()
        print(f"Error en rechazo: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ==========================================================
# INSCRIPCIÓN DE ASIGNATURAS
# ==========================================================
@app.route('/api/inscribir', methods=['POST'])
@requiere_rol('estudiante')
def api_inscribir_asignatura():
    data = request.get_json()
    asignatura_id = data.get('asignatura_id')

    estudiante = Estudiante.query.filter_by(usuario_id=session['usuario_id']).first()
    if not estudiante:
        return jsonify({'ok': False, 'msg': 'Estudiante no encontrado'}), 404

    existe = EstudianteAsignatura.query.filter_by(
        estudiante_id=estudiante.id,
        asignatura_id=asignatura_id
    ).first()
    if existe:
        return jsonify({'ok': False, 'msg': 'Ya inscrito'}), 400

    inscripcion = EstudianteAsignatura()
    inscripcion.estudiante_id = estudiante.id
    inscripcion.asignatura_id = asignatura_id
    db.session.add(inscripcion)
    db.session.commit()

    return jsonify({'ok': True, 'msg': 'Inscripción exitosa'})


# ==========================================================
# CERRAR SESION
# ==========================================================

@app.route('/logout')
@login_required
def logout():
    # 1. Avisar a Flask-Login que cierre la sesión
    logout_user()
    
    # 2. Limpiar todos los datos de la sesión de Flask
    session.clear()
    
    # 3. Crear la respuesta de redirección
    response = make_response(redirect(url_for('login_page')))
    
    # 4. BORRADO MANUAL DE COOKIES
    # Flask suele usar 'session' o 'remember_token'
    response.delete_cookie('session')
    response.delete_cookie('remember_token')
    
    # 5. Forzar al navegador a no usar caché (para que refresque el menú)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    flash("Has salido del sistema.", "info")
    return response


# ==========================================================
# API: MENSAJERÍA (Mensajes)
# ==========================================================
@app.route('/mensajeria') # Ruta base
@app.route('/mensajeria/<int:receptor_id>') # Ruta con ID
@login_required
def bandeja_entrada(receptor_id=None):
    # Detectar qué pestaña quiere ver el usuario (por defecto 'recibidos')
    vista = request.args.get('vista', 'recibidos')
    
    # Si receptor_id es None, intentamos buscarlo en los parámetros ?id=XX (por si acaso)
    if receptor_id is None:
        receptor_id = request.args.get('id', type=int)

    # Consultas para los contadores
    total_recibidos = Mensaje.query.filter_by(receptor_id=current_user.id, leido=False).count()
    total_enviados = Mensaje.query.filter_by(emisor_id=current_user.id, enviado=True).count()
    total_favoritos = Mensaje.query.filter(
        db.or_(Mensaje.receptor_id == current_user.id, Mensaje.emisor_id == current_user.id),
        Mensaje.mensaje_favorito == True
    ).count()

    # Lógica para llenar la lista principal
    if vista == 'enviados':
        mensajes_mostrar = Mensaje.query.filter_by(emisor_id=current_user.id, enviado=True).order_by(Mensaje.fecha.desc()).all()
    elif vista == 'favoritos':
        mensajes_mostrar = Mensaje.query.filter(
            db.or_(Mensaje.receptor_id == current_user.id, Mensaje.emisor_id == current_user.id),
            Mensaje.mensaje_favorito == True
        ).order_by(Mensaje.fecha.desc()).all()
    else:
        # Recibidos
        mensajes_mostrar = Mensaje.query.filter_by(receptor_id=current_user.id, recibido=True).order_by(Mensaje.fecha.desc()).all()
    
    # Lista de usuarios para el buscador del modal
    usuarios = Usuario.query.filter(Usuario.rol != 'admin', Usuario.id != current_user.id).all()
    
    # LÓGICA DE PRESELECCIÓN: Buscamos al usuario si hay un receptor_id
    usuario_preseleccionado = None
    if receptor_id and receptor_id != current_user.id:
        usuario_preseleccionado = Usuario.query.get(receptor_id)

    return render_template('mensajeria.html', 
                           mensajes=mensajes_mostrar, 
                           total_recibidos=total_recibidos,
                           total_enviados=total_enviados,
                           total_favoritos=total_favoritos,
                           usuarios=usuarios, 
                           vista_actual=vista,
                           hoy=datetime.utcnow().date(),
                           body_class='fondo-mensajes',
                           usuario_preseleccionado=usuario_preseleccionado)

# SESIÓN DE MANEJO ENVÍO DE MENSAJES
@app.route('/enviar_mensaje', methods=['POST'])
@login_required
def enviar_mensaje():
    email_dest = request.form.get('email_institucional')
    contenido = request.form.get('contenido')
    archivo = request.files.get('archivo_adjunto')

    if not email_dest or not contenido:
        flash("El destinatario y el mensaje son obligatorios.", "warning")
        return redirect(url_for('bandeja_entrada'))

    receptor = Usuario.query.filter_by(correo_institucional=email_dest).first()

    if not receptor:
        flash("Error: El destinatario no existe en el sistema.", "danger")
        return redirect(url_for('bandeja_entrada'))

    nombre_archivo = None
    if archivo and archivo.filename != '':
        nombre_archivo = secure_filename(f"msg_{datetime.now().timestamp()}_{archivo.filename}")
        archivo.save(os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo))

    try:
        nuevo_mensaje = Mensaje(
            emisor_id=current_user.id,
            receptor_id=receptor.id,
            contenido=contenido,
            archivo_adjunto=nombre_archivo,
            enviado=True,
            recibido=True,
            leido=False
        )
        db.session.add(nuevo_mensaje)
        db.session.flush()

        nueva_notif = Notificacion(
            usuario_id=receptor.id,
            tipo='mensaje',
            mensaje=f"Nuevo mensaje de {current_user.nombre}",
            leida=False,
            item_id=nuevo_mensaje.id
        )
        db.session.add(nueva_notif)
        db.session.commit()

        flash(f"Mensaje enviado a {receptor.nombre} correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error interno al procesar el envío.", "danger")

    return redirect(url_for('bandeja_entrada', vista='enviados'))

@app.route('/ver_mensaje/<int:id>')
@login_required
def ver_mensaje(id):
    mensaje = Mensaje.query.get_or_404(id)
    
    if mensaje.receptor_id != current_user.id and mensaje.emisor_id != current_user.id:
        flash("No tienes permiso para acceder a este mensaje.", "danger")
        return redirect(url_for('bandeja_entrada'))

    if mensaje.receptor_id == current_user.id:
        if not mensaje.leido:
            mensaje.leido = True
        
        notif = Notificacion.query.filter_by(
            usuario_id=current_user.id, 
            item_id=mensaje.id, 
            tipo='mensaje'
        ).first()
        
        if notif:
            notif.leida = True
            
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()

    return render_template('ver_detalle_mensaje.html', 
                           msg=mensaje, 
                           hoy=datetime.utcnow().date())

# ELIMINAR MENSAJE
@app.route('/eliminar_mensaje/<int:id>', methods=['POST'])
@login_required
def eliminar_mensaje(id):
    mensaje = Mensaje.query.get_or_404(id)
    
    # Seguridad: Solo el emisor o receptor pueden borrarlo
    if mensaje.emisor_id == current_user.id or mensaje.receptor_id == current_user.id:
        try:
            db.session.delete(mensaje)
            db.session.commit()
            flash("Mensaje eliminado correctamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash("Error al eliminar el mensaje.", "danger")
    
    return redirect(url_for('bandeja_entrada', receptor_id=usuario.id))



# ==========================
# RUTA PERFIL
# ==========================

# Guardar imagen de perfil
def save_profile_image(foto):
    ext = foto.filename.rsplit('.', 1)[-1].lower()
    if ext not in FileAllowed_PERFILES_EXIT:
        return None
    nombre_archivo = f"{uuid.uuid4().hex}_{secure_filename(foto.filename)}"
    carpeta = os.path.join(app.static_folder, 'uploads', 'perfiles')
    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(carpeta, nombre_archivo)
    foto.save(ruta)
    return nombre_archivo

# Guardar imagen de portada
def save_cover_image(foto):
    ext = foto.filename.rsplit('.', 1)[-1].lower()
    if ext not in FileAllowed_PERFILES_EXIT:
        return None
    nombre_archivo = f"{uuid.uuid4().hex}_{secure_filename(foto.filename)}"
    carpeta = os.path.join(app.static_folder, 'uploads', 'portadas')
    os.makedirs(carpeta, exist_ok=True)
    ruta = os.path.join(carpeta, nombre_archivo)
    foto.save(ruta)
    return nombre_archivo


# Ver perfil
@app.route('/perfil/<int:user_id>')
def ver_perfil(user_id):
    usuario = Usuario.query.get_or_404(user_id)
    if not usuario:
        flash("Usuario no encontrado", "danger")
        return redirect(url_for('inicio'))

    # 2. ¿Es mi propio perfil? (Para saber si mostramos botones de editar)
    es_mio = (user_id == current_user.id)

    # Limitar a las 5 noticias más recientes
    if usuario.noticias:
        usuario.noticias_ultimas = sorted(usuario.noticias, key=lambda x: x.fecha, reverse=True)[:5]
    else:
        usuario.noticias_ultimas = []

    return render_template('perfil.html', usuario=usuario, puedo_editar=es_mio)



# RUTA: EDITAR PERFIL
from sqlalchemy.orm.attributes import flag_modified

@app.route('/perfil/editar', methods=['GET', 'POST'])
@login_required
def perfil_editar():
    usuario = current_user 

    if request.method == 'POST':
        usuario.curso = request.form.get('curso')

        # Foto de perfil (esto funciona bien porque es un string)
        foto = request.files.get('foto_perfil')
        if foto:
            filename = save_profile_image(foto)
            if filename: 
                usuario.foto_perfil = filename

        # --- LOGICA PARA PICKLETYPE (PORTADAS) ---
        fotos_portada = request.files.getlist('foto_portada')
        
        # Solo procesamos si el usuario seleccionó archivos nuevos
        if fotos_portada and fotos_portada[0].filename != '':
            nuevas_fotos = [] # Creamos una lista limpia
            
            for una_foto in fotos_portada[:3]:
                filename_p = save_cover_image(una_foto)
                if filename_p:
                    nuevas_fotos.append(filename_p)
            
            # ASIGNACIÓN CRÍTICA
            usuario.foto_portada = nuevas_fotos 
            
            # Obligamos a la DB a reconocer el cambio en el campo binario
            flag_modified(usuario, "foto_portada")

        try:
            db.session.commit()
            flash("Perfil actualizado correctamente", "success")
        except Exception as e:
            db.session.rollback()
            print(f"Error al guardar: {e}")
            flash("Error al guardar los cambios", "danger")

        return redirect(url_for('ver_perfil', user_id=usuario.id))

    return render_template('perfil_editar.html', usuario=usuario)

# RUTA: CAMBIAR CONTRASEÑA
@app.route('/perfil/cambiar_contrasena', methods=['GET', 'POST'])
@login_required
def perfil_cambiar_contrasena():
    usuario = current_user
    if request.method == 'POST':
        actual = request.form.get('actual')
        nueva = request.form.get('nueva')
        confirmar = request.form.get('confirmar')

        if not check_password_hash(usuario.password_hash, actual):
            flash("Contraseña actual incorrecta", "danger")
            return redirect(url_for('perfil_cambiar_contrasena'))

        if nueva != confirmar:
            flash("Las contraseñas no coinciden", "danger")
            return redirect(url_for('perfil_cambiar_contrasena'))

        usuario.password_hash = generate_password_hash(nueva)
        db.session.commit()
        flash("Contraseña actualizada", "success")
        return redirect(url_for('ver_perfil', user_id=current_user.id))

    return render_template('perfil_cambiar_contrasena.html')


# RUTAS DEL PERFIL
@app.route('/configuracion')
def configuracion():
    return render_template('configuracion.html')

@app.route('/estudios')
def estudios():
    return render_template('estudios.html')



# ==========================
# NOTICIAS
# ==========================
# Lista de etiquetas permitidas
tags_permitidas = ['p','b','i','u','ul','li','a','img','strong','em','br','h1','h2','h3']

# Atributos permitidos
atributos_permitidos = {
    'a': ['href', 'title', 'target'],
    'img': ['src', 'alt', 'title', 'width', 'height'],
}


# CREAR NUEVA NOTICIA
@app.route('/nueva_noticia', methods=['GET', 'POST'])
@login_required # Usamos el decorador oficial de Flask-Login
def nueva_noticia():
    form = NoticiaForm()
    
    # Ya no necesitamos buscar autor_id en session, usamos current_user
    if form.validate_on_submit():
        
        # 1. El ID viene directamente de Flask-Login
        autor_id = current_user.id 
        
        # 2. Manejo de archivos (Imagen/Video)
        archivo = form.archivo.data
        nombre_archivo = None
        tipo_archivo = None

        if archivo:
            nombre_archivo = secure_filename(archivo.filename)
            archivo.save(os.path.join(UPLOAD_FOLDER, nombre_archivo))
            ext = nombre_archivo.rsplit('.', 1)[1].lower()

            if ext in ALLOWED_IMAGE_EXT:
                tipo_archivo = 'imagen'
            elif ext in ALLOWED_VIDEO_EXT:
                tipo_archivo = 'video'
            else:
                flash('Formato de imagen/video no permitido', 'danger')
                return redirect(url_for('nueva_noticia'))
            
        # 3. Guardar documento Word o PDF
        documento = form.documento.data
        nombre_documento = None
        if documento:
            nombre_documento = secure_filename(documento.filename)
            documento.save(os.path.join(UPLOAD_FOLDER, nombre_documento))

        # 4. Crear noticia usando current_user.id
        # 4. Crear noticia usando current_user.id
        try:
            noticia = Noticia(
                titulo=form.titulo.data,
                contenido=form.contenido.data,
                fecha=date.today(),
                archivo=nombre_archivo,
                tipo_archivo=tipo_archivo,
                destacado=form.destacado.data,
                documento=nombre_documento,
                autor_id=autor_id,
                pie_archivo=form.pie_archivo.data
            )

            db.session.add(noticia)
            
            # --- CAMBIO CRÍTICO AQUÍ ---
            # Esto le asigna un ID a la noticia en la base de datos sin cerrar la transacción
            db.session.flush() 

            # 3. LÓGICA DE NOTIFICACIÓN
            usuarios_comunidad = Usuario.query.filter(
                Usuario.id != current_user.id, 
                Usuario.rol != 'admin'
            ).all()

            for usuario in usuarios_comunidad:
                notificacion = Notificacion(
                    usuario_id=usuario.id,
                    tipo='noticia',
                    item_id=noticia.id, # Ahora noticia.id YA tiene valor gracias al flush()
                    mensaje=f"🔔 {current_user.nombre} publicó una nueva noticia: {noticia.titulo}",
                    leida=False,
                    fecha_creacion=datetime.utcnow()
                )
                db.session.add(notificacion)

            db.session.commit()
            flash('Noticia publicada correctamente', 'success')
            return redirect(url_for('noticia_completa', noticia_id=noticia.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error en la base de datos: {str(e)}', 'danger')
            return redirect(url_for('nueva_noticia'))

    # Para el GET, también pasamos current_user.id al template si es necesario
    return render_template('nueva_noticia.html', form=form, autor_id=current_user.id)



# Noticia completa
@app.route('/noticias/<int:noticia_id>')
def noticia_completa(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    return render_template('noticia_completa.html', noticia=noticia)


# Lista de todas las noticias
@app.route('/lista_noticias')
def lista_noticias():
    noticias = Noticia.query.order_by(Noticia.fecha.desc()).all()
    return render_template('lista_noticias.html', noticias=noticias)


# Eliminar noticia
@app.route('/eliminar_noticia/<int:noticia_id>', methods=['POST'])
@login_required
def eliminar_noticia(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    
    # Doble seguro en el servidor
    if not (noticia.autor_id == current_user.id or current_user.rol == 'admin'):
        flash("Acceso denegado.", "danger")
        return redirect(url_for('lista_noticias'))

    db.session.delete(noticia)
    db.session.commit()
    flash("Noticia eliminada correctamente.", "success")
    return redirect(url_for('lista_noticias'))

# Descargar documento asociado a noticia
@app.route('/descargar_noticia/<int:noticia_id>')
@requiere_login
def descargar_noticia(noticia_id):
    noticia = Noticia.query.get_or_404(noticia_id)
    if not noticia.documento:
        flash('No hay documento disponible', 'warning')
        return redirect(url_for('noticias_page'))

    return send_file(
        os.path.join(UPLOAD_FOLDER, noticia.documento),
        as_attachment=True
    )

#Ordenar noticas por fecha descendente.Mostrar solo hasta 9 noticias
@app.route('/noticias')
def noticias_destacadas():
    noticias = Noticia.query.order_by(Noticia.fecha.desc()).limit(9).all()
    return render_template("noticias.html", noticias=noticias)


# PODER COMENTAR LAS NOTICIAS
@app.route('/noticia/<int:noticia_id>/comentar', methods=['POST'])
@login_required
def agregar_comentario_noticia(noticia_id):
    contenido = request.form.get('contenido')
    if not contenido:
        flash("El comentario no puede estar vacío.", "warning")
        return redirect(url_for('noticia_completa', noticia_id=noticia_id))

    nuevo_comentario = Comentario(
        contenido=contenido,
        noticia_id=noticia_id,
        autor_id=current_user.id
    )
    
    db.session.add(nuevo_comentario)
    db.session.commit()
    flash("Comentario publicado correctamente.", "success")
    return redirect(url_for('noticia_completa', noticia_id=noticia_id))


# eliminar un comenario
@app.route('/eliminar_comentario/<int:comentario_id>', methods=['POST'])
@login_required
def eliminar_comentario_noticia(comentario_id):
    comentario = Comentario.query.get_or_404(comentario_id)
    noticia_id = comentario.noticia_id # Guardamos el ID para poder regresar
    
    # Verificación de seguridad: Solo autor del comentario o admin
    if not (comentario.autor_id == current_user.id or current_user.rol == 'admin'):
        flash("No tienes permiso para eliminar este comentario.", "danger")
        return redirect(url_for('noticia_completa', noticia_id=noticia_id))

    try:
        db.session.delete(comentario)
        db.session.commit()
        flash("Comentario eliminado.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error al eliminar el comentario.", "danger")

    return redirect(url_for('noticia_completa', noticia_id=noticia_id))


# ==========================================================
# BIBLIOTECA
# ==========================================================

# Método para guardar un PDF
def save_pdf_file(f):
    if not f or getattr(f, 'filename', '') == '':
        return None

    # Nombre seguro
    filename = secure_filename(f.filename)
    ext = filename.rsplit('.', 1)[-1].lower()

    # Verificar extensión
    if ext != 'pdf':
        return None

    # Crear nombre único
    unique_name = f"{uuid.uuid4().hex}_{filename}"

    # Carpeta destino: static/libros/
    carpeta_libros = os.path.join(app.root_path, "static/uploads/libros")
    os.makedirs(carpeta_libros, exist_ok=True)

    # Guardar archivo
    f.save(os.path.join(carpeta_libros, unique_name))

    return unique_name



# Método para guardar portadas de documentos
def save_portada_file(file):
    if not file or file.filename == "":
        return None

    filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
    path = os.path.join(app.root_path, "static/uploads/portadas", filename)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    file.save(path)

    return filename


# ------------------------------
# RUTAS
# ------------------------------

# LIBROS FÍSICOS
@app.route('/biblioteca/fisicos')
@requiere_login
def biblioteca_fisicos():
    libros = Biblioteca.query.filter_by(tipo_libro='fisico').all()
    return render_template('libros_fisicos.html', libros_fisicos=libros, body_class='libros-fisicos')


# LIBROS DIGITALES
@app.route('/biblioteca/digitales')
@requiere_login
def biblioteca_digitales():

    libros = Biblioteca.query.filter_by(tipo_libro='libro').all()

    # Filtro por titulación solo para TFG
    filtro_titulacion = request.args.get("titulacion", None)

    query_tfg = Biblioteca.query.filter_by(tipo_libro='tfg')

    if filtro_titulacion:
        query_tfg = query_tfg.filter_by(titulacion=filtro_titulacion)

    tfg_publicos = query_tfg.all()

    return render_template(
        'libros_digitales.html',
        libros=libros,
        tfg_publicos=tfg_publicos,
        filtro_titulacion=filtro_titulacion,
        body_class="libros-digitales"
    )


# ELIMINAR libro o TFG
@app.route('/biblioteca/eliminar/<int:item_id>', methods=['POST'])
@requiere_login
def biblioteca_eliminar(item_id):
    item = Biblioteca.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for('biblioteca_digitales'))


# CREAR o EDITAR LIBRO/TFG
@app.route('/biblioteca/editar', methods=['GET', 'POST'])
@app.route('/biblioteca/editar/<int:item_id>', methods=['GET', 'POST'])
@requiere_login
def biblioteca_editar(item_id=None):

    item = Biblioteca.query.get(item_id) if item_id else None
    form = BibliotecaForm(obj=item)

    if item:
        form.publico.data = bool(item.publico)

    if form.validate_on_submit():

        # ------------------------------
        # GUARDAR PORTADA
        # ------------------------------
        portada_file = form.portada.data
        portada_filename = None

        if portada_file and portada_file.filename != "":
            portada_filename = save_portada_file(portada_file)

        # ------------------------------
        # PDF O LINK
        # ------------------------------
        tipo = form.tipo.data   # pdf o link
        filename = None
        enlace = None

        if tipo == 'pdf':
            archivo_pdf = form.archivo_pdf.data
            if archivo_pdf:
                filename = save_pdf_file(archivo_pdf)

            if not filename and not (item and item.archivo):
                flash("Debes subir un PDF válido.", "warning")
                return redirect(request.url)

        elif tipo == 'link':
            enlace = form.enlace.data
            if not enlace:
                flash("Debes introducir un enlace válido.", "warning")
                return redirect(request.url)

        # ------------------------------
        # ACTUALIZAR
        # ------------------------------
        if item:

            item.titulo = form.titulo.data
            item.descripcion = form.descripcion.data
            item.tipo = tipo
            item.enlace = enlace
            item.tipo_libro = form.tipo_libro.data
            item.publico = bool(form.publico.data)

            # Solo TFG lleva titulación
            item.titulacion = form.titulacion.data if item.tipo_libro == "tfg" else None

            if filename:
                item.archivo = filename
            if portada_filename:
                item.portada = portada_filename

        # ------------------------------
        # CREAR NUEVO
        # ------------------------------
        else:
            nuevo = Biblioteca(
                titulo=form.titulo.data,
                descripcion=form.descripcion.data,
                tipo=tipo,
                archivo=filename,
                enlace=enlace,
                tipo_libro=form.tipo_libro.data,
                publico=bool(form.publico.data),
                usuario_id=session['usuario_id'],
                portada=portada_filename,
                titulacion=form.titulacion.data if form.tipo_libro.data == "tfg" else None
            )

            db.session.add(nuevo)

        db.session.commit()
        flash("Elemento guardado correctamente.", "success")

        return redirect(url_for('biblioteca_digitales'))

    return render_template('biblioteca_editar.html', form=form, item=item, body_class="libros-digitales")


# Ruta libro en físico


# Agregar libro físico
@app.route('/biblioteca/fisicos/agregar', methods=['GET', 'POST'])
@requiere_login
def agregar_libro_fisico():
    form = LibroFisicoForm()
    
    if form.validate_on_submit():
        # Guardar portada si se sube
        portada_file = form.portada.data
        portada_filename = None
        if portada_file and portada_file.filename != "":
            nombre_original = secure_filename(portada_file.filename)
            portada_filename = f"{uuid.uuid4().hex}_{nombre_original}"
            
            # Carpeta de portadas para libros físicos
            carpeta_portadas = os.path.join(app.root_path, "static/uploads/libros")
            os.makedirs(carpeta_portadas, exist_ok=True)
            
            # Guardar archivo en disco
            portada_file.save(os.path.join(carpeta_portadas, portada_filename))
        
        # Crear nuevo libro físico
        nuevo_libro = Biblioteca(
            titulo=form.titulo.data,
            descripcion=form.descripcion.data,
            tipo='fisico',          # Obligatorio para la DB
            tipo_libro='fisico',    # Marca que es libro físico
            portada=portada_filename, # Nombre de archivo
            usuario_id=session['usuario_id'],
            publico=True             # Opcional, por defecto público
        )

        db.session.add(nuevo_libro)
        db.session.commit()
        flash("Libro físico agregado correctamente.", "success")
        return redirect(url_for('biblioteca_fisicos'))

    return render_template('agregar_libro_fisico.html', form=form, body_class="libros-fisicos")


# Eliminar libro físico
@app.route('/biblioteca/fisicos/eliminar/<int:item_id>', methods=['POST'])
@requiere_login
def eliminar_libro_fisico(item_id):
    libro = Biblioteca.query.get_or_404(item_id)
    # Opcional: eliminar archivo de portada del disco
    if libro.portada:
        ruta_portada = os.path.join(app.root_path, "static/uploads/libros", libro.portada)
        if os.path.exists(ruta_portada):
            os.remove(ruta_portada)
    db.session.delete(libro)
    db.session.commit()
    flash("Libro físico eliminado.", "success")
    return redirect(url_for('biblioteca_page'))

# Editar libro fisico
# Editar libro físico
@app.route('/biblioteca/fisicos/editar/<int:item_id>', methods=['GET', 'POST'])
@requiere_login
def editar_libro_fisico(item_id):
    libro = Biblioteca.query.get_or_404(item_id)
    form = LibroFisicoForm(obj=libro)  # Carga los datos actuales en el formulario

    if form.validate_on_submit():
        # Actualizar título y descripción
        libro.titulo = form.titulo.data
        libro.descripcion = form.descripcion.data

        # Guardar nueva portada si se sube
        portada_file = form.portada.data
        if portada_file and portada_file.filename != "":
            # Eliminar portada anterior del disco (opcional)
            if libro.portada:
                ruta_portada_ant = os.path.join(app.root_path, "static/uploads/libros", libro.portada)
                if os.path.exists(ruta_portada_ant):
                    os.remove(ruta_portada_ant)

            # Guardar nueva portada
            nombre_original = secure_filename(portada_file.filename)
            portada_filename = f"{uuid.uuid4().hex}_{nombre_original}"
            carpeta_portadas = os.path.join(app.root_path, "static/uploads/libros")
            os.makedirs(carpeta_portadas, exist_ok=True)
            portada_file.save(os.path.join(carpeta_portadas, portada_filename))
            libro.portada = portada_filename

        db.session.commit()
        flash("Libro físico actualizado correctamente.", "success")
        return redirect(url_for('biblioteca_fisicos'))

    return render_template('agregar_libro_fisico.html', form=form, editar=True)



# SOLICITAR PRESTAMO DEL LIBRP
@app.route('/biblioteca/fisicos/prestamo/<int:id>', methods=['GET', 'POST'])
@requiere_login
def solicitar_prestamo(id):
    libro = Biblioteca.query.get_or_404(id)
    form = SolicitudPrestamoForm()

    if form.validate_on_submit():
        # Datos del usuario desde la sesión
        nombre = session.get('nombre')
        apellidos = session.get('apellidos')
        dip = session.get('dip')
        correo = session.get('correo')
        fecha_envio = datetime.utcnow()
        motivo = form.motivo.data

        msg = Message(
            subject=f"Solicitud de préstamo: {libro.titulo}",
            sender=correo,
            recipients=["bibliotecario@universidad.com"],  # correo del bibliotecario
            body=f"""
Solicitud de préstamo de libro:

Libro: {libro.titulo}
Solicitante: {nombre} {apellidos}
DIP: {dip}
Correo: {correo}
Fecha: {fecha_envio.strftime('%d/%m/%Y %H:%M')}
Motivo: {motivo}
"""
        )
        mail.send(msg)

        flash("Solicitud enviada correctamente.", "success")
        return redirect(url_for('biblioteca_fisicos'))

    return render_template('solicitud_prestamo.html', form=form, libro=libro, body_class="prestar-libro")




# =========================================================
# SELECTIVIDAD
# =========================================================

# 1. LISTADO GENERAL (Vista de cuadrícula tipo noticias)
@app.route('/selectividad')
def selectividad():
    # Buscamos todos los registros
    datos = Selectividad.query.order_by(Selectividad.fecha_publicacion.desc()).all()
    
    # Pasamos 'resultados' al HTML para que el bucle {% for res in resultados %} funcione
    return render_template('selectividad_listado.html', resultados=datos)

# 2. DETALLE DE UNA NOTICIA (Vista individual al hacer clic)
@app.route('/selectividad/<int:id>', methods=['GET', 'POST'])
def selectividad_detalle(id):
    # Buscamos la noticia específica por ID
    resultado = Selectividad.query.get_or_404(id)
    form_opinion = OpinionForm()

    # Lógica para recibir comentarios en la noticia
    if form_opinion.validate_on_submit():
        nueva_opinion = OpinionSelectividad(
            nombre_usuario=form_opinion.nombre.data,
            comentario=form_opinion.mensaje.data,
            selectividad_id=id
        )
        db.session.add(nueva_opinion)
        db.session.commit()
        flash("Tu opinión ha sido publicada.", "success")
        return redirect(url_for('selectividad_detalle', id=id))

    return render_template('selectividad_detalle.html', resultado=resultado, form=form_opinion)

# 3. FORMULARIO DE SUBIDA (Solo administrador)
# 3. FORMULARIO DE SUBIDA (Actualizado para CKEditor)
@app.route('/subir-selectividad', methods=['GET', 'POST'])
@requiere_login
@requiere_rol(['directivo', 'admin']) # Corrección de la lógica de roles
def subir_selectividad():
    form = SelectividadForm()
    
    if form.validate_on_submit():
        try:
            # 1. Guardar el PDF (Obligatorio)
            archivo_pdf = form.pdf_file.data
            nombre_pdf = guardar_archivo(archivo_pdf, 'pdfs_selectividad') 
            
            # 2. Guardar la Foto (Opcional)
            nombre_foto = None
            if form.foto_examen.data:
                nombre_foto = guardar_archivo(form.foto_examen.data, 'fotos_selectividad')

            # 3. Crear el registro en la DB
            # Nota: comentario_admin ahora guarda etiquetas HTML (<b>, <p>, etc.)
            nueva_entrada = Selectividad(
                titulo=form.titulo.data.strip(),
                comentario_admin=form.comentario_admin.data, # CKEditor envía HTML aquí
                ruta_pdf=nombre_pdf,
                ruta_foto=nombre_foto,
                ruta_pie_foto=form.pie_foto.data,
                fecha_publicacion=datetime.utcnow()
            )
            
            db.session.add(nueva_entrada)
            db.session.commit()
            
            flash("Resultados publicados con éxito en el portal oficial.", "success")
            return redirect(url_for('selectividad'))

        except Exception as e:
            db.session.rollback()
            print(f"Error en la subida: {e}")
            flash("Hubo un error al procesar los archivos. Inténtelo de nuevo.", "danger")
    else:
            print(f"Errores de WTForms: {form.errors}")

    return render_template('subir_selectividad.html', form=form)

# 4. ELIMINAR OPINIÓN (Solo administrador)
@app.route('/eliminar-opinion/<int:id>')
@requiere_login
@requiere_rol(['admin', 'directivo'])
def eliminar_opinion(id):
    opinion = OpinionSelectividad.query.get_or_404(id)
    id_noticia = opinion.selectividad_id
    db.session.delete(opinion)
    db.session.commit()
    flash("Comentario eliminado correctamente.", "warning")
    return redirect(url_for('selectividad_detalle', id=id_noticia))



# ELIMINAR NOTICIA DE SELECTIVIDAD
@app.route('/eliminar-selectividad/<int:id>', methods=['POST'])
@requiere_login
@requiere_rol(['admin', 'administrador', 'directivo'])
def eliminar_selectividad(id):
    noticia = Selectividad.query.get_or_404(id)
    
    try:
        # OPCIONAL: Borrar archivos físicos del servidor para liberar espacio
        if noticia.ruta_pdf:
            path_pdf = os.path.join(app.root_path, 'static', noticia.ruta_pdf)
            if os.path.exists(path_pdf): os.remove(path_pdf)
            
        if noticia.ruta_foto:
            path_foto = os.path.join(app.root_path, 'static', noticia.ruta_foto)
            if os.path.exists(path_foto): os.remove(path_foto)

        # Borrar registro de la base de datos
        db.session.delete(noticia)
        db.session.commit()
        
        flash("La noticia ha sido eliminada permanentemente.", "info")
    except Exception as e:
        db.session.rollback()
        print(f"Error al eliminar noticia: {e}")
        flash("Error técnico al intentar eliminar la noticia.", "danger")
        
    return redirect(url_for('selectividad'))

# =========================================================
# SOLICITUD DE MATRICULA. ESTUDIANTES NUEVOS, EGRESADOS, CONTINUANTES
# =========================================================
@app.route('/solicitar-matricula', methods=['GET', 'POST'])
def solicitar_matricula():
    form = MatriculaForm()
    
    if form.validate_on_submit():
        try:
            # A. Procesar archivos dinámicamente
            files_data = {}
            file_fields = [
                'doc_dni', 'doc_cert_selectividad', 'doc_instancia', 'doc_hoja_bachillerato',
                'doc_foto_carnet', 'doc_conducta_comunidad', 'doc_conducta_centro',
                'doc_ficha_matricula', 'doc_ficha_permanencia', 'doc_hoja_facultad',
                'doc_acta_defensa', 'doc_convalidaciones', 'doc_homologacion'
            ]

            for field in file_fields:
                file_storage = getattr(form, field).data
                if file_storage:
                    files_data[field] = guardar_archivo(file_storage, 'matriculas_docs')
                else:
                    files_data[field] = None

            # B. Crear registro en BD
            nueva_solicitud = SolicitudMatricula(
                tipo_estudiante=form.tipo_estudiante.data,
                nombre=form.nombre.data,
                apellidos=form.apellidos.data,
                fecha_nacimiento=form.fecha_nacimiento.data,
                residencia=form.residencia.data,
                natural_de=form.natural_de.data,
                dni_numero=form.dni_numero.data,
                distrito_provincia=form.distrito_provincia.data,
                email=form.email.data,
                carrera=form.carrera.data,
                telefono=form.telefono.data,
                sexo=form.sexo.data,           
                nacionalidad=form.nacionalidad.data,
                **files_data # Pasa todos los archivos del diccionario de golpe
                )
                
            db.session.add(nueva_solicitud)
            db.session.commit()

            # CORREO AUTOMATICO QUE LLEGA TRAS COMPLETAR LA SOLICITUD::
            enviar_acuse_recibo(solicitud=nueva_solicitud, dominio=request.host_url.rstrip('/'))
        

            # Redirigir a Éxito
            return redirect(url_for('pago_exitoso', 
                                    nombre=form.nombre.data, 
                                    apellidos=form.apellidos.data, 
                                    carrera=form.carrera.data))

        except Exception as e:
            db.session.rollback()
            return redirect(url_for('pago_error', mensaje=f"Error en base de datos: {str(e)}"))

    # Manejo de errores de validación (Si falta un campo obligatorio)
    if request.method == 'POST':
        errores = [f"{getattr(form, f).label.text}: {m[0]}" for f, m in form.errors.items()]
        return redirect(url_for('pago_error', mensaje=" | ".join(errores)))

    return render_template('solicitar_matricula.html', form=form)


# RUTA DE ÉXITO
@app.route('/inscripcion-exitosa')
def pago_exitoso(): # <--- Este es el 'endpoint'
    datos = {
        'nombre': request.args.get('nombre'),
        'apellidos': request.args.get('apellidos'),
        'carrera': request.args.get('carrera')
    }
    return render_template('inscripcion_ok.html', datos=datos)

# RUTA DE ERROR
@app.route('/error-inscripcion')
def pago_error(): # <--- Asegúrate de que se llame exactamente así
    mensaje = request.args.get('mensaje', 'Error desconocido en el formulario')
    return render_template('inscripcion_error.html', mensaje=mensaje)


# 1. LISTADO GENERAL
@app.route('/admin/ver-matriculas')
def ver_matriculas():
    solicitudes = SolicitudMatricula.query.order_by(SolicitudMatricula.fecha_creacion.desc()).all()
    return render_template('admin_matriculas.html', solicitudes=solicitudes)


# ADMITIR ALUMNO
@app.route('/admin/matricula/estado/<int:id>/Admitido')
def admitir_alumno(id):
    solicitud = SolicitudMatricula.query.get_or_404(id)
    solicitud.estado = 'Admitido'
    db.session.commit()

    # Definimos el dominio aquí
    dominio = "http://localhost:5000"
    
    # Pasamos el dominio a la función (asegúrate de que la función lo acepte)
    try:
        enviar_confirmacion_matricula(solicitud, dominio)
        flash("Alumno admitido y notificado", "success")
    except Exception as e:
        flash(f"Error al enviar correo: {str(e)}", "warning")
    
    return redirect(url_for('ver_matriculas'))


# SOLICITAR REVISION.
@app.route('/admin/matricula/estado/<int:id>/Revision')
def solicitar_revision(id):
    # 1. Buscamos la solicitud en la base de datos
    solicitud = SolicitudMatricula.query.get_or_404(id)
    
    # 2. Actualizamos el estado
    solicitud.estado = 'Revision'
    db.session.commit()

    # 3. LLAMADA AL NUEVO AVISO:
    # Pasamos el objeto 'solicitud' completo, el dominio automático y un comentario opcional
    enviar_aviso_revision(
        solicitud=solicitud, 
        dominio=request.host_url.rstrip('/'),
        comentario="Algunos documentos son ilegibles o están incompletos. Por favor, revísalos y vuelve a subirlos."
    )
    
    flash(f"La solicitud #{id} ha sido marcada para revisión y se ha enviado el correo al alumno.", "warning")
    return redirect(url_for('ver_matriculas'))


# PERMITIR DESCARGAR PDF SI EL ALUMNO ES ADMITIDO
@app.route('/descargar-admision/<int:id>')
def descargar_admision(id):
    # 1. Buscamos la solicitud en la base de datos
    solicitud = SolicitudMatricula.query.get_or_404(id)
    
    # 2. Verificación de seguridad: Solo admitidos pueden descargar
    if solicitud.estado != 'Admitido':
        # Si intenta descargar sin estar admitido, lanzamos error 403 (Prohibido)
        return """
        <div style="text-align:center; margin-top:50px; font-family:Arial;">
            <h1 style="color:red;">Acceso Denegado</h1>
            <p>Tu solicitud aún está en proceso de revisión o no ha sido admitida.</p>
            <a href="/">Volver al inicio</a>
        </div>
        """, 403

    try:
        # 3. Llamamos a la función que creamos antes (que devuelve el BytesIO)
        pdf_buffer = generar_pdf_admision(solicitud)
        
        # 4. Enviamos el archivo al navegador
        # as_attachment=True fuerza la descarga
        # download_name es el nombre que verá el alumno al guardar el archivo
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Resguardo_Admision_{solicitud.nombre}_{solicitud.id}.pdf"
        )

    except Exception as e:
        # En caso de error técnico (ej. falta una librería), lo registramos
        logging.error(f"Error generando PDF para ID {id}: {e}")
        return f"Hubo un error al generar tu certificado: {str(e)}", 500

# 3. EXPORTACIÓN A EXCEL
@app.route('/admin/exportar-matriculas')
@requiere_login
@requiere_rol('administrador')
def exportar_matriculas():
    solicitudes = SolicitudMatricula.query.all()
    data = [{
        'Fecha': s.fecha_creacion.strftime('%d/%m/%Y'),
        'Estudiante': f"{s.nombre} {s.apellidos}",
        'DNI': s.dni_numero,
        'Carrera': s.carrera,
        'Estado': s.estado
    } for s in solicitudes]
    
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Matriculas_UNGE.xlsx')


# ========================================================
# CORREOS DE LA UNIVERSIDAD PARA LA MATRICULA
# ========================================================

# Estructura de un mensaje si la matricula es aceptada. GMAIL/OUTLOOK
def enviar_confirmacion_matricula(solicitud, dominio):
    """
    Envía un correo de admisión con diseño institucional.
    Usa el objeto 'solicitud' para obtener los datos reales de la BD.
    """
    # Extraemos los datos del objeto solicitud
    email_destino = solicitud.email
    nombre_alumno = f"{solicitud.nombre} {solicitud.apellidos}"
    id_solicitud = solicitud.id
    dni_alumno = solicitud.dni_numero
    carrera_alumno = solicitud.carrera

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Secretaría Académica UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = f"¡FELICIDADES! Has sido admitido en la UNGE - Ref: {id_solicitud}"

    # Construcción del HTML con un bloque de mensaje elegante
    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Facultad de Ciencias de la Salud</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #1a237e;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: 300;">¡ADMISIÓN CONFIRMADA!</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a <strong>{nombre_alumno}</strong>,</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Es un placer para nosotros informarle que, tras la revisión exhaustiva de su expediente académico y documentación presentada, 
                        ha sido formalmente <strong>ACEPTADO</strong> para cursar estudios en nuestra institución.
                    </p>

                    <div style="background-color: #f8fafb; border-radius: 10px; padding: 25px; border-left: 5px solid #ff6f00; margin-bottom: 30px;">
                        <table width="100%" style="font-size: 14px; color: #37474f;">
                            <tr>
                                <td style="padding-bottom: 10px;"><strong>Nº DE EXPEDIENTE:</strong></td>
                                <td style="padding-bottom: 10px; text-align: right;">#{id_solicitud}</td>
                            </tr>
                            <tr>
                                <td style="padding-bottom: 10px;"><strong>TIPO DE MATRÍCULA:</strong></td>
                                <td style="padding-bottom: 10px; text-align: right; color: #ff6f00;">{solicitud.tipo_estudiante}</td>
                            </tr>
                            <tr>
                                <td style="padding-bottom: 10px;"><strong>DOCUMENTO IDENTIDAD:</strong></td>
                                <td style="padding-bottom: 10px; text-align: right;">{dni_alumno}</td>
                            </tr>
                            <tr>
                                <td><strong>CARRERA ASIGNADA:</strong></td>
                                <td style="text-align: right; color: #1a237e; font-weight: bold;">{carrera_alumno}</td>
                            </tr>
                        </table>
                    </div>

                    <p style="font-size: 14px; color: #455a64; margin-bottom: 30px; text-align: center;">
                        Para completar su proceso de matriculación, debe descargar su <strong>Resguardo de Admisión</strong> y presentarlo en la secretaría Edi. II de la Facultad para el pago de tasas.
                    </p>

                    <div style="text-align: center; margin: 20px 0 40px 0;">
                        <a href="{dominio}/descargar-admision/{id_solicitud}" 
                           style="background-color: #ff6f00; color: #ffffff; padding: 18px 35px; text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 15px; display: inline-block; box-shadow: 0 4px 6px rgba(255,111,0,0.2);">
                           OBTENER MI RESGUARDO DE ADMISIÓN
                        </a>
                    </div>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Secretaría Académica - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar correo: {e}")
        return False

# --------------------------------------------------------------------------
# Mensaje automatico que se recibe tras solicitar matricula
def enviar_acuse_recibo(solicitud, dominio):
    """
    Envía un correo de confirmación de recepción con el diseño institucional.
    Usa el objeto 'solicitud' para mantener la coherencia con enviar_confirmacion_matricula.
    """
    email_destino = solicitud.email
    nombre_alumno = f"{solicitud.nombre} {solicitud.apellidos}"
    id_solicitud = solicitud.id
    carrera_alumno = solicitud.carrera

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Secretaría Académica UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = f"Solicitud de Matrícula Recibida - Ref: {id_solicitud}"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Facultad de Ciencias de la Salud</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #1a237e;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 300; letter-spacing: 1px;">SOLICITUD RECIBIDA</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a <strong>{nombre_alumno}</strong>,</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Le confirmamos que hemos recibido correctamente sus datos y documentos para la inscripción en la carrera de <strong>{carrera_alumno}</strong>. 
                        Su solicitud ha entrado en fase de revisión por parte de la Secretaría Académica.
                    </p>

                    <div style="background-color: #f8fafb; border-radius: 10px; padding: 25px; border-left: 5px solid #1a237e; margin-bottom: 30px;">
                        <table width="100%" style="font-size: 14px; color: #37474f;">
                            <tr>
                                <td style="padding-bottom: 10px;"><strong>Nº DE REFERENCIA:</strong></td>
                                <td style="padding-bottom: 10px; text-align: right; font-weight: bold;">#{id_solicitud}</td>
                            </tr>
                            <tr>
                                <td style="padding-bottom: 10px;"><strong>ESTADO ACTUAL:</strong></td>
                                <td style="padding-bottom: 10px; text-align: right; color: #ff6f00; font-weight: bold;">En Revisión</td>
                            </tr>
                            <tr>
                                <td><strong>CARRERA SOLICITADA:</strong></td>
                                <td style="text-align: right; color: #1a237e;">{carrera_alumno}</td>
                            </tr>
                        </table>
                    </div>

                    <p style="font-size: 14px; color: #455a64; margin-bottom: 30px; line-height: 1.6;">
                        <strong>¿Qué sigue ahora?</strong><br>
                        Nuestro equipo verificará que toda la documentación cargada sea legible y válida. Una vez finalizada la revisión, recibirá un nuevo correo electrónico indicando si su solicitud ha sido <strong>admitida</strong> o si requiere alguna corrección.
                    </p>

                    <p style="font-size: 13px; color: #78909c; text-align: center; font-style: italic;">
                        No es necesario que realice ninguna otra acción por el momento.
                    </p>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Secretaría Académica - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Este es un mensaje automático, por favor no responda a este correo.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar acuse de recibo: {e}")
        return False

# -----------------------------------------------------------------
# GENERAR EN DOCUMENTOS DESCARGABLES LA INFORMACION DE LOS USUARIOS
def generar_pdf_admision(solicitud):
    # Usamos FPDF en modo estándar
    pdf = FPDF()
    pdf.add_page()

    # --- INSERTAR LOGO ---
    # Buscamos la ruta absoluta de la imagen en tu carpeta static
    ruta_logo = os.path.join(current_app.root_path, 'static', 'img', 'logo_unge.jpeg')
    
    try:
        # image(ruta, x, y, ancho)
        pdf.image(ruta_logo, 10, 8, 25) 
    except Exception as e:
        print(f"No se pudo cargar el logo: {e}")
    
    # Encabezado UNGE
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt="UNIVERSIDAD NACIONAL DE GUINEA ECUATORIAL", ln=True, align='C')
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(200, 10, txt="FACULTAD DE CIENCIAS DE LA SALUD", ln=True, align='C')
    pdf.ln(10)
    
    # Título del documento
    pdf.set_fill_color(230, 230, 230)
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 12, txt="RESGUARDO DE ADMISIÓN", ln=True, align='C', fill=True)
    pdf.ln(10)
    
    # Datos del Alumno - Limpiamos caracteres extraños con .encode().decode() si es necesario
    pdf.set_font("Arial", '', 12)
    pdf.cell(0, 10, txt=f"ID de Solicitud: #00{solicitud.id}", ln=True)
    pdf.cell(0, 10, txt=f"Nombre Completo: {solicitud.nombre} {solicitud.apellidos}", ln=True)
    pdf.cell(0, 10, txt=f"DNI / Pasaporte: {solicitud.dni_numero}", ln=True)
    pdf.set_font("Arial", 'B', 12)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, txt=f"Tipo de Estudiante: {solicitud.tipo_estudiante.upper()}", ln=True)
    pdf.cell(0, 10, txt=f"Carrera: {solicitud.carrera}", ln=True)
    pdf.ln(5)
    
    # Estado
    pdf.set_text_color(0, 128, 0) # Verde institucional
    pdf.cell(0, 10, txt="ESTADO: ADMITIDO / APROBADO", ln=True)
    pdf.set_text_color(0, 0, 0)
    
    pdf.ln(15)
    pdf.set_font("Arial", 'I', 10)
    text_footer = ("Este documento acredita que el estudiante ha sido aceptado oficialmente. "
                   "Debe presentarse en la Secretaría de la Facultad para formalizar el pago.")
    pdf.multi_cell(0, 10, txt=text_footer)
    
    pdf.ln(20)
    # Fecha y Firma
    pdf.cell(0, 10, f"Fecha de emision: {datetime.now().strftime('%d/%m/%Y')}", ln=True, align='R')
    pdf.ln(20)
    pdf.cell(0, 10, "__________________________", ln=True, align='C')
    pdf.cell(0, 7, "Sello y Firma de Secretaria Academica", ln=True, align='C')

    # EXPLICACIÓN DEL FIX:
    # 1. Obtenemos el output como string 'S'
    # 2. Lo codificamos a latin-1 ignorando caracteres que FPDF no soporte
    # 3. Lo metemos en BytesIO para que send_file lo vea como un ARCHIVO BINARIO
    output = pdf.output(dest='S').encode('latin-1', 'ignore')
    buffer = io.BytesIO(output)
    buffer.seek(0) # Ponemos el puntero al inicio para que Flask lea desde el principio
    
    return buffer


#-------------------------------------------------------------------
# FUNCION PARA LA DOCUMENTACION INCOMPLETA. MATRICULA-SECRETARIA
def enviar_aviso_revision(solicitud, dominio, comentario="No especificado"):
    """
    Envía un correo informando que la documentación es incorrecta.
    Permite incluir una nota del administrador explicando el error.
    """
    email_destino = solicitud.email
    nombre_alumno = f"{solicitud.nombre} {solicitud.apellidos}"
    id_solicitud = solicitud.id

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Secretaría Académica UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = f"ACCIÓN REQUERIDA: Documentación Incompleta - Ref: {id_solicitud}"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Facultad de Ciencias de la Salud</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #ff6f00;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 300; letter-spacing: 1px;">DOCUMENTACIÓN PENDIENTE</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Hola <strong>{nombre_alumno}</strong>,</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Hemos revisado su solicitud <strong>#{id_solicitud}</strong> y lamentamos informarle que su expediente está 
                        <span style="color: #d32f2f; font-weight: bold;">INCOMPLETO</span> o contiene documentos que no han podido ser validados.
                    </p>

                    <div style="background-color: #fff3e0; border-radius: 10px; padding: 25px; border-left: 5px solid #ff6f00; margin-bottom: 30px;">
                        <p style="margin: 0 0 10px 0; font-size: 14px; color: #e65100; font-weight: bold;">OBSERVACIONES DE SECRETARÍA:</p>
                        <p style="margin: 0; font-size: 15px; color: #3e2723; font-style: italic;">
                            "{comentario}"
                        </p>
                    </div>

                    <p style="font-size: 14px; color: #455a64; margin-bottom: 30px; line-height: 1.6;">
                        <strong>¿Cómo solucionar esto?</strong><br>
                        Debe acudir a la Secretaría de la Facultad o volver a realizar el proceso de carga de documentos asegurándose de que los archivos sean legibles, estén en formato PDF y correspondan a lo solicitado.
                    </p>

                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{dominio}/solicitar-matricula" 
                           style="background-color: #1a237e; color: #ffffff; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 14px; display: inline-block;">
                            VOLVER AL FORMULARIO DE MATRÍCULA
                        </a>
                    </div>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Secretaría Académica - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Si tiene dudas, por favor contacte con nosotros directamente.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar aviso de revisión: {e}")
        return False
# ----------------------------------------------------------------


# ==========================================================
# EXPEDIENTE ACADEMICO
# ==========================================================
# RUTA 1: Ver el expediente y los cálculos
# Eliminamos la ruta doble y usamos una sola que sea clara
@app.route('/expediente/<int:user_id>')
@login_required
def ver_expediente(user_id):
    # 1. Buscamos al dueño del expediente
    usuario_perfil = Usuario.query.get_or_404(user_id)
    
    # 2. Buscamos (o creamos) su registro en la tabla Estudiante
    estudiante = Estudiante.query.filter_by(usuario_id=user_id).first()
    
    if not estudiante:
        estudiante = Estudiante(
            usuario_id=user_id, 
            matricula=f"MAT-{user_id}", 
            carrera=usuario_perfil.carrera 
        )
        db.session.add(estudiante)
        db.session.commit()

    # 3. Traemos sus notas
    registros = Expediente.query.filter_by(estudiante_id=estudiante.id).all()
    
    # 4. Cálculos
    puntuacion = sum(r.nota_final for r in registros)
    total_materias = len(registros)
    promedio = round(puntuacion / total_materias, 2) if total_materias > 0 else 0

    # 5. LÓGICA DE VISUALIZACIÓN (La clave de Facebook)
    # Si el ID de la URL es igual al ID del que está logueado, puede editar
    puedo_editar = (current_user.id == user_id)

    return render_template(
        'expediente.html', 
        usuario=usuario_perfil, 
        estudiante=estudiante, 
        registros=registros, 
        puntuacion=puntuacion, 
        aprobado=promedio, 
        puedo_editar=puedo_editar
    )

# RUTA 2: Procesar la firma de la nota
# RUTA 2: Procesar la firma de la nota
@app.route('/firmar-nota/<int:nota_id>')
@login_required
def firmar_nota(nota_id):
    nota = Expediente.query.get_or_404(nota_id)
    estudiante_propietario = Estudiante.query.get(nota.estudiante_id)

    # Verificación de seguridad
    if estudiante_propietario.usuario_id != current_user.id:
        flash("Acción denegada: No puedes firmar documentos ajenos.", "danger")
        # Ajustamos 'id' por 'user_id' para que coincida con la ruta del expediente
        return redirect(url_for('ver_expediente', user_id=current_user.id))

    # Si todo es correcto, firmamos
    if not nota.firmado:
        try:
            nota.firmado = True
            nota.fecha_firma = datetime.now()
            db.session.commit()
            flash("Nota firmada digitalmente con éxito.", "success")
        except Exception as e:
            db.session.rollback()
            flash("Error al procesar la firma electrónica.", "danger")
    
    # Redirigimos al expediente del usuario actual usando 'user_id'
    return redirect(url_for('ver_expediente', user_id=current_user.id))

# ENVIAR NOTAS DESDE EXCEL CORREGIDO
@app.route('/importar-notas', methods=['POST'])
@login_required
def importar_notas():
    carrera_modal = request.form.get('carrera_nombre')
    materia_modal = request.form.get('materia_nombre')
    semestre_modal = request.form.get('semestre_nombre') 
    
    # --- GENERACIÓN DINÁMICA DEL AÑO ACADÉMICO ---
    anio_actual = datetime.now().year
    anio_siguiente = anio_actual + 1
    rango_academico = f"{anio_actual}-{anio_siguiente}" 
    
    archivo = request.files.get('archivo_excel')
    
    # Validación de seguridad por si no suben archivo
    if not archivo:
        flash("No se seleccionó ningún archivo.", "warning")
        return redirect(url_for('profesor_panel', user_id=current_user.id))

    try:
        df = pd.read_excel(archivo)
        df.columns = [c.strip().capitalize() for c in df.columns] 

        for index, row in df.iterrows():
            dip_excel = str(row['Dip']).split('.')[0].strip()
            nota_excel = row['Nota']

            usuario = Usuario.query.filter_by(
                dip=dip_excel, 
                carrera=carrera_modal, 
                rol='estudiante'
            ).first()

            if usuario:
                est = Estudiante.query.filter_by(usuario_id=usuario.id).first()
                
                if est:
                    nueva_nota = Expediente(
                        estudiante_id=est.id,
                        asignatura_nombre=materia_modal, 
                        semestre=semestre_modal,
                        nota_final=nota_excel,
                        anio_academico=rango_academico,
                        firmado=False
                    )
                    db.session.add(nueva_nota)
        
        db.session.commit()
        flash(f"Notas de {materia_modal} ({rango_academico}) cargadas con éxito.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error al procesar el Excel o guardar: {str(e)}", "danger")

    # LA CORRECCIÓN CLAVE:
    # Ahora pasamos el user_id del profesor actual para que Flask sepa a dónde volver
    return redirect(url_for('profesor_panel', user_id=current_user.id))

# ==========================================================
# PANEL OFICIAL PARA PROFESORES
# ==========================================================
# Un estudiante no puede entrar
@app.errorhandler(403)
def access_denied(error):
    return render_template('errors/403.html'), 403



# SUBIR NOTAS EN EXEL
@app.route('/subir_notas', methods=['POST'])
@login_required
def subir_notas():
    if 'archivo_excel' not in request.files:
        flash("No se seleccionó ningún archivo", "warning")
        return redirect(request.url)
    
    file = request.files['archivo_excel']
    if file.filename == '':
        return redirect(request.url)

    if file:
        df = pd.read_excel(file)
        
        # Ejemplo de estructura esperada en Excel: 
        # Columnas: [matricula, asignatura, nota, anio]
        
        for index, row in df.iterrows():
            # Buscamos al estudiante por matrícula
            estudiante = Estudiante.query.filter_by(matricula=row['matricula']).first()
            
            if estudiante:
                nuevo_registro = Expediente(
                    estudiante_id=estudiante.id,
                    asignatura_nombre=row['asignatura'],
                    nota_final=row['nota'],
                    anio_academico=row['anio'],
                    firmado=True,
                    fecha_firma=datetime.utcnow()
                )
                db.session.add(nuevo_registro)
        
        db.session.commit()
        flash("Notas procesadas y publicadas correctamente", "success")
        
    return redirect(url_for('panel_profesor'))



# REGISTRO DE PROFESORES
# ==========================================================
# RUTAS DEL PROFESOR (Registro, Activación y Panel)
# ==========================================================

# Definimos las variables de ruta (igual que hacías antes)
UPLOAD_FOLDER_DIP = os.path.join('static', 'uploads', 'dips')
UPLOAD_FOLDER_FOTOS = os.path.join('static', 'uploads', 'fotos')

# 1. REGISTRO DEL PROFESOR
@app.route('/profesor/registro', methods=['GET', 'POST'])
def profesor_registro():
    if request.method == 'POST':
        # 1. Captura de datos
        email_personal = request.form.get('email_personal')
        dip_numero = request.form.get('dip_numero')
        
        # 2. Validaciones (Aspirante o Usuario existente)
        aspirante_existente = Profesor.query.filter_by(dip_aspirante=dip_numero).first()
        usuario_existente = Usuario.query.filter_by(dip=dip_numero).first()

        if aspirante_existente or usuario_existente:
            flash(f"El DIP {dip_numero} ya está registrado o tiene una solicitud pendiente.", "warning")
            return redirect(url_for('profesor_registro'))

        # 3. Manejo de Archivos
        dip_file = request.files.get('archivo_dip')
        foto_file = request.files.get('archivo_foto')
        
        if dip_file and foto_file:
            # --- CREACIÓN AUTOMÁTICA DE CARPETAS ---
            # Esto es lo que permite que se creen solas si no existen
            os.makedirs(UPLOAD_FOLDER_DIP, exist_ok=True)
            os.makedirs(UPLOAD_FOLDER_FOTOS, exist_ok=True)
            # ---------------------------------------

            nombre_dip = secure_filename(f"dip_{dip_numero}_{dip_file.filename}")
            nombre_foto = secure_filename(f"foto_{dip_numero}_{foto_file.filename}")
            
            # Guardamos usando las variables locales (SIN app.config)
            dip_file.save(os.path.join(UPLOAD_FOLDER_DIP, nombre_dip))
            foto_file.save(os.path.join(UPLOAD_FOLDER_FOTOS, nombre_foto))

            # 4. Crear el registro en la tabla Profesor (Aspirante)
            nuevo_aspirante = Profesor(
                nombre_aspirante=request.form.get('nombre'),
                apellidos_aspirante=request.form.get('apellidos'),
                correo_personal=email_personal,
                dip_aspirante=dip_numero,
                telefono_aspirante=request.form.get('telefono'),
                sexo_aspirante=request.form.get('sexo'),
                departamento=request.form.get('departamento'),
                especialidad=request.form.get('especialidad'),
                archivo_dip=nombre_dip,
                archivo_foto=nombre_foto,
                cuenta_activa=False 
            )
            
            try:
                db.session.add(nuevo_aspirante)
                db.session.commit()
                
                # Guardamos el ID en sesión para la página de pendiente
                session['aspirante_reciente_id'] = nuevo_aspirante.id
                
                enviar_correo_recepcion_profesor(nuevo_aspirante, request.host_url)
                
                flash("Solicitud enviada con éxito.", "success")
                return redirect(url_for('profesor_registro_pendiente'))
                
            except Exception as e:
                db.session.rollback()
                flash(f"Error al guardar: {str(e)}", "danger")
                return redirect(url_for('profesor_registro'))

    return render_template('profesor_registro.html')

# 2. MENSAJE QUE MUESTRA EL ENVIO DE FORMULARIO
@app.route('/profesor/registro/pendiente')
def profesor_registro_pendiente():
    # 1. Intentamos obtener el ID del ASPIRANTE (Profesor) guardado en el paso anterior
    aspirante_id = session.get('aspirante_reciente_id')
    
    if not aspirante_id:
        # Si no hay ID en sesión, redirigimos al inicio
        return redirect(url_for('index'))
    
    # 2. Buscamos en la tabla Profesor
    aspirante = Profesor.query.get(aspirante_id)
    
    if not aspirante:
        return redirect(url_for('index'))
    
    # 3. Pasamos el objeto 'aspirante' al HTML
    return render_template('profesor_registro_pendiente.html', aspirante=aspirante)

# 3. LUGAR PARA PODER ACTIVAR LA CUENTA
@app.route('/profesor/activar', methods=['GET', 'POST'])
def profesor_activar_cuenta():
    if request.method == 'POST':
        codigo_ingresado = request.form.get('codigo_activacion')
        nueva_password = request.form.get('nueva_password')
        confirmar_password = request.form.get('confirmar_password')
        
        # 1. Buscar al profesor por el código
        profe = Profesor.query.filter_by(codigo_activacion=codigo_ingresado).first()
        
        if not profe:
            flash("El código introducido no es válido o ya fue utilizado.", "danger")
            return redirect(url_for('profesor_activar_cuenta'))

        # 2. Validar que las contraseñas coincidan
        if nueva_password != confirmar_password:
            flash("Las contraseñas no coinciden. Inténtelo de nuevo.", "warning")
            return render_template('profesor_activar_cuenta.html', codigo=codigo_ingresado)

        # 3. Validar longitud mínima de contraseña
        if len(nueva_password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "warning")
            return render_template('profesor_activar_cuenta.html', codigo=codigo_ingresado)

        try:
            # 4. Actualizar el estado del Profesor
            profe.cuenta_activa = True
            profe.codigo_activacion = None  # El código queda inutilizable tras el éxito
            
            # 5. Actualizar la contraseña en la tabla Usuario vinculada
            usuario_vinculado = profe.usuario
            usuario_vinculado.set_password(nueva_password)
            
            db.session.commit()
            
            # Pasamos el objeto usuario para mostrar el correo institucional en el éxito
            return render_template('profesor_activacion_exito.html', usuario=usuario_vinculado)
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error al activar la cuenta: {str(e)}", "danger")
            
    return render_template('profesor_activar_cuenta.html')

# PANEL PRINCIPAL DEL PROFESOR (PERFIL)
@app.route('/profesor/panel/<int:user_id>')
@login_required
def profesor_panel(user_id):
    # 1. Buscamos al dueño del perfil (el ID que viene en la URL)
    usuario_profe = Usuario.query.get_or_404(user_id)
    
    # 2. Buscamos sus datos profesionales en la tabla Profesor
    # IMPORTANTE: Buscamos por el user_id de la URL, no por el current_user
    datos_profe = Profesor.query.filter_by(usuario_id=user_id).first()
    
    # 3. LÓGICA TIPO FACEBOOK:
    # ¿El que está logueado es el mismo que el dueño del perfil?
    puedo_editar = (current_user.id == user_id)
    
    # 4. Seguridad: Si el usuario de la URL no es realmente un profesor, 
    # evitamos que entre a esta plantilla
    if usuario_profe.rol != 'profesor':
        flash("Este perfil no pertenece a un docente.", "info")
        return redirect(url_for('facultad'))

    # Si el profesor no tiene datos creados aún, pero es el dueño,
    # podrías crearlos aquí o simplemente mostrar un mensaje
    if not datos_profe and puedo_editar:
        # Aquí podrías crear un perfil básico de profesor si no existe
        pass

    return render_template('profesor_panel.html', 
                           usuario=usuario_profe, 
                           profe=datos_profe, 
                           puedo_editar=puedo_editar)


# LOS PROFES PUEDEN GUARDAR DOCUMENTOS
@app.route('/guardar-documento-profesor', methods=['POST'])
@login_required
@requiere_rol(['profesor', 'admin', 'administrador'])
def guardar_documento_profesor():
    # 1. Obtener al profesor vinculado al usuario actual
    profesor = Profesor.query.filter_by(usuario_id=current_user.id).first()
    if not profesor:
        flash("Perfil de profesor no encontrado", "danger")
        return redirect(url_for('inicio'))

    archivo = request.files.get('archivo')
    tipo = request.form.get('tipo') # Viene del select: 'actas' o 'examenes'

    if archivo and tipo in ['actas', 'examenes']:
        # 2. Guardar archivo físico
        filename = secure_filename(f"{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{archivo.filename}")
        
        # Definimos la ruta de la carpeta
        folder_path = os.path.join(app.root_path, 'static', 'uploads', 'recursos')
        # --- SOLUCIÓN: Crear la carpeta si no existe ---
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        upload_path = os.path.join(folder_path, filename)
        archivo.save(upload_path)
        
        # 3. Actualizar el JSON
        # Convertimos el string de la DB a un diccionario de Python
        docs = json.loads(profesor.mis_documentos_json) if profesor.mis_documentos_json else {"actas": [], "examenes": []}
        
        # Añadimos la nueva ruta y el nombre
        docs[tipo].append({
            "nombre": archivo.filename,
            "ruta": filename,
            "fecha": datetime.now().strftime('%d/%m/%Y %H:%M')
        })

        # Convertimos de nuevo a string para guardar en la DB
        profesor.mis_documentos_json = json.dumps(docs)
        db.session.commit()

        flash(f"{tipo.capitalize()} guardado correctamente", "success")
    
    return redirect(url_for('mis_documentoss_vista')) # Nombre de tu ruta de vista

# Ver los documemtos guardados

@app.route('/profesor/mis-documentos')
@login_required
@requiere_rol(['profesor', 'admin', 'administrador'])
def mis_documentos_vista():
    # Buscamos al profesor vinculado al usuario logueado
    profesor = Profesor.query.filter_by(usuario_id=current_user.id).first()
    
    if not profesor:
        flash("No se encontró perfil de profesor vinculado.", "warning")
        return redirect(url_for('inicio'))

    # Convertimos el string JSON de la DB a un diccionario de Python
    # Si está vacío, usamos la estructura base
    try:
        documentos = json.loads(profesor.mis_documentos_json) if profesor.mis_documentos_json else {"actas": [], "examenes": []}
    except:
        documentos = {"actas": [], "examenes": []}
    
    return render_template('profesor_mis_documentos.html', documentos=documentos)

# POder eliminar un documento
@app.route('/eliminar-documento/<string:tipo>/<string:filename>')
@login_required
@requiere_rol(['profesor', 'admin', 'administrador'])
def eliminar_documento(tipo, filename):
    profesor = Profesor.query.filter_by(usuario_id=current_user.id).first()
    if not profesor or not profesor.mis_documentos_json:
        flash("No se pudo encontrar el documento.", "danger")
        return redirect(url_for('mis_documentos_vista'))

    # 1. Cargar el JSON actual
    docs = json.loads(profesor.mis_documentos_json)

    if tipo in docs:
        # 2. Filtrar la lista para quitar el archivo que coincida con el nombre físico
        # Guardamos la lista vieja para comparar si realmente se borró algo
        lista_original = docs[tipo]
        docs[tipo] = [d for d in docs[tipo] if d['ruta'] != filename]

        if len(docs[tipo]) < len(lista_original):
            # 3. Borrar el archivo físico del servidor
            file_path = os.path.join(app.root_path, 'static', 'uploads', 'recursos', filename)
            if os.path.exists(file_path):
                os.remove(file_path)

            # 4. Guardar el nuevo JSON en la base de datos
            profesor.mis_documentos_json = json.dumps(docs)
            db.session.commit()
            flash("Documento eliminado correctamente.", "info")
        else:
            flash("No se encontró el archivo en tus registros.", "warning")

    return redirect(url_for('mis_documentos_vista'))

# ENTREGAR O ENVIAR ACTAS A LA SECRETARIA
@app.route('/entregar-acta-oficial', methods=['GET', 'POST'])
@login_required
def profesor_entregar_acta():
    # Buscamos al profesor vinculado al usuario actual
    profesor = Profesor.query.filter_by(usuario_id=current_user.id).first()
    
    if request.method == 'POST':
        # Capturamos los datos del formulario (Selects)
        titulacion = request.form.get('titulacion')
        periodo = request.form.get('periodo')
        curso = request.form.get('curso')
        asignatura = request.form.get('asignatura')
        archivo = request.files.get('archivo')

        if archivo and titulacion:
            # 1. Gestión del Archivo Físico
            ext = archivo.filename.rsplit('.', 1)[1].lower()
            # Nombre profesional: ACTA_Carrera_ID_Fecha.ext
            safe_titulacion = secure_filename(titulacion.replace(" ", "_"))
            filename = f"ACTA_{safe_titulacion}_{current_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
            
            folder = os.path.join(app.root_path, 'static', 'uploads', 'entregas')
            if not os.path.exists(folder): 
                os.makedirs(folder)
                
            archivo.save(os.path.join(folder, filename))

            # 2. Gestión de la Base de Datos (JSON)
            # Cargamos el JSON existente o creamos uno nuevo si está vacío
            if profesor.entregas_oficiales_json:
                data = json.loads(profesor.entregas_oficiales_json)
            else:
                data = {"entregas": []}
            
            # Construcción del objeto de entrega con datos automáticos
            nueva_entrega = {
                "id_entrega": int(datetime.now().timestamp()),
                "titulacion": titulacion,
                "periodo": periodo,
                "curso": curso,
                "asignatura": asignatura,
                "profesor_nombre": f"{current_user.nombre} {current_user.apellidos}", # Automático
                "archivo": filename,
                "fecha": datetime.now().strftime('%d/%m/%Y %H:%M'), # Automático
                "estado": "Enviado",
                "observaciones": ""
            }
            
            # Guardamos y confirmamos en DB
            data["entregas"].append(nueva_entrega)
            profesor.entregas_oficiales_json = json.dumps(data)
            db.session.commit()
            
            flash("Acta enviada oficialmente con éxito", "success")
            return redirect(url_for('profesor_entregar_acta'))

    # Cargar historial para enviarlo al HTML (GET)
    try:
        historial = json.loads(profesor.entregas_oficiales_json) if profesor.entregas_oficiales_json else {"entregas": []}
    except:
        historial = {"entregas": []}
        
    return render_template('profesor_entregar_acta.html', entregas=historial["entregas"])


# LA SECRETARIA O EL ADMINISTRADOR PUEDEN VER LAS ACTAS ENVIADAS
@app.route('/secretaria/bandeja-actas')
@login_required
@requiere_rol(['admin', 'profesor']) # Solo personal autorizado
def secretaria_actas():
    todos_los_profesores = Profesor.query.all()
    todas_las_entregas = []

    for profe in todos_los_profesores:
        if profe.entregas_oficiales_json:
            data = json.loads(profe.entregas_oficiales_json)
            for entrega in data.get("entregas", []):
                # Añadimos el ID del profesor para poder identificar de quién es cada acta
                entrega['usuario_id'] = profe.usuario_id 
                todas_las_entregas.append(entrega)

    # Ordenar por fecha (descendente) usando el timestamp que creamos
    todas_las_entregas.sort(key=lambda x: x.get('id_entrega', 0), reverse=True)

    return render_template('secretaria_actas.html', entregas=todas_las_entregas)

# Poder cambia el estado de entrega
@app.route('/secretaria/cambiar-estado-acta/<int:u_id>/<e_id>/<string:nuevo_estado>')
@login_required
# Asegúrate de agregar 'secretaria' o 'secretario' según tu DB
@requiere_rol(['profesor', 'admin']) 
def cambiar_estado_acta(u_id, e_id, nuevo_estado):
    # Buscamos al profesor por su usuario_id
    profesor = Profesor.query.filter_by(usuario_id=u_id).first()
    
    if profesor and profesor.entregas_oficiales_json:
        try:
            data = json.loads(profesor.entregas_oficiales_json)
            encontrado = False
            
            for entrega in data.get("entregas", []):
                # Convertimos ambos a string para una comparación segura
                if str(entrega.get("id_entrega")) == str(e_id):
                    entrega["estado"] = nuevo_estado
                    encontrado = True
                    break
            
            if encontrado:
                # Marcamos el campo como modificado para SQLAlchemy
                profesor.entregas_oficiales_json = json.dumps(data)
                db.session.commit()
                flash(f"Estado actualizado a: {nuevo_estado}", "success")
            else:
                flash("No se encontró la entrega específica.", "warning")
                
        except Exception as e:
            db.session.rollback()
            flash(f"Error al procesar: {str(e)}", "danger")
    
    # Redirigir a la bandeja de entrada de secretaría
    return redirect(url_for('secretaria_actas'))

# Poder eliminar una entrega
@app.route('/secretaria/eliminar-acta-oficial/<int:u_id>/<float:e_id>')
@login_required
@requiere_rol(['admin', 'profesor'])
def eliminar_acta_notas(u_id, e_id):
    # 1. Buscar al profesor dueño del acta
    profesor = Profesor.query.filter_by(usuario_id=u_id).first()
    
    if profesor and profesor.entregas_oficiales_json:
        data = json.loads(profesor.entregas_oficiales_json)
        
        # 2. Encontrar el archivo antes de filtrar para borrarlo físicamente
        acta_a_borrar = next((e for e in data["entregas"] if e["id_entrega"] == e_id), None)
        
        if acta_a_borrar:
            # Borrado físico del archivo
            file_path = os.path.join(app.root_path, 'static', 'uploads', 'entregas', acta_a_borrar["archivo"])
            if os.path.exists(file_path):
                os.remove(file_path)
            
            # 3. Filtrar el JSON para quitar la entrega (se borra para el profesor también)
            data["entregas"] = [e for e in data["entregas"] if e["id_entrega"] != e_id]
            
            # 4. Guardar cambios
            profesor.entregas_oficiales_json = json.dumps(data)
            db.session.commit()
            flash("Acta eliminada permanentemente del sistema.", "danger")
        else:
            flash("No se encontró el registro del acta.", "warning")
            
    return redirect(url_for('secretaria_actas'))

# EDITAR EL PERFIL DEL PROFESOR
@app.route('/perfil-profesores/editar', methods=['GET', 'POST'])
@login_required
@requiere_rol(['profesor'])
def profesor_editar_perfil():
    # Buscamos al profesor vinculado al usuario logueado
    profe = Profesor.query.filter_by(usuario_id=current_user.id).first()
    
    if request.method == 'POST':
        # 1. Actualizar campos de texto
        profe.departamento = request.form.get('departamento')
        profe.especialidad = request.form.get('especialidad')
        profe.titulo_academico = request.form.get('titulo_academico')
        
        # 2. Manejar la Foto de Perfil
        file = request.files.get('foto')
        if file and file.filename != '':
            ext = file.filename.rsplit('.', 1)[1].lower()
            if ext in ['jpg', 'jpeg', 'png']:
                filename = secure_filename(f"FOTO_{current_user.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
                folder = os.path.join(app.root_path, 'static/uploads/fotos')
                if not os.path.exists(folder): os.makedirs(folder)
                
                # Borrar foto anterior si existe para no llenar el servidor
                if profe.archivo_foto:
                    old_path = os.path.join(folder, profe.archivo_foto)
                    if os.path.exists(old_path): os.remove(old_path)
                
                file.save(os.path.join(folder, filename))
                profe.archivo_foto = filename

        db.session.commit()
        flash("Perfil actualizado con éxito", "success")
        return redirect(url_for('profesor_panel', user_id=current_user.id))

    # 'usuario' es current_user, 'puedo_editar' es True ya que es su propio perfil
    return render_template('profesor_editar_perfil.html', 
                           usuario=current_user, 
                           profe=profe, 
                           puedo_editar=True)

# ==========================================================
# RUTAS DE ADMINISTRADOR (Gestión y Validación de Docentes)
# ==========================================================

@app.route('/admin/profesores/revision')
@login_required
def admin_profesor_revision():
    """Panel principal donde el admin ve a los profes que esperan validación"""
    if current_user.rol != 'admin':
        flash("Acceso denegado. Solo administradores.", "danger")
        return redirect(url_for('inicio'))
    
    # Solo traemos a los profesores cuya cuenta_activa es False
    solicitudes = Profesor.query.filter_by(cuenta_activa=False).all()
    return render_template('admin_profesor_revision.html', solicitudes=solicitudes)

# Validar al profesor
@app.route('/admin/profesores/validar/<int:id>', methods=['POST'])
@login_required
@requiere_rol("admin")
def admin_profesor_validar(id):
    profe = Profesor.query.get_or_404(id)
    
    try:
        # 1. GENERAR DATOS PRIMERO (Evita inconsistencias)
        # Generar Correo Institucional
        nombres = profe.nombre_aspirante.lower().split()
        iniciales = "".join([n[0] for n in nombres if n])
        primer_apellido = profe.apellidos_aspirante.lower().split()[0]
        correo_inst = f"{iniciales}.{primer_apellido}.2025@profesores.cienciasdelasalud.unge.gq"
        
        # Generar Código de Activación
        codigo = "".join(random.choices(string.ascii_uppercase + string.digits, k=8))

        # 2. CREAR USUARIO CON LOS DATOS FINALES
        nuevo_usuario = Usuario(
            nombre=profe.nombre_aspirante,
            apellidos=profe.apellidos_aspirante,
            correo=profe.correo_personal,
            correo_institucional=correo_inst,  # <--- LOGIN OFICIAL
            rol='profesor',
            dip=profe.dip_aspirante,
            telefono=profe.telefono_aspirante,
            sexo=profe.sexo_aspirante
        )
        nuevo_usuario.set_password(profe.dip_aspirante) 
        
        db.session.add(nuevo_usuario)
        db.session.flush() 

        # 3. VINCULAR Y ACTUALIZAR TABLA PROFESOR
        profe.usuario_id = nuevo_usuario.id 
        profe.codigo_activacion = codigo
        profe.cuenta_activa = False # Se activa cuando el use el código
        
        db.session.commit()
        
        # 4. ENVIAR CORREO (Asegúrate que enviar_activacion_profesor acepte estos 4 datos)
        enviar_activacion_profesor(profe, codigo, correo_inst, request.host_url)
        
        flash(f"Validación exitosa. Se ha enviado el código al correo personal del profesor.", "success")
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR EN VALIDACIÓN: {e}") # Importante revisar tu terminal
        flash(f"Error al validar: {str(e)}", "danger")
        
    return redirect(url_for('admin_profesor_revision'))


@app.route('/admin/profesores/rechazar/<int:id>', methods=['POST'])
@login_required
@requiere_rol("admin")
def admin_profesor_rechazar(id):
    """Rechaza la solicitud y borra los datos para permitir re-intento"""
    if current_user.rol != 'admin': return abort(403)
    
    profe = Profesor.query.get_or_404(id)
    usuario = profe.usuario
    motivo = request.form.get('motivo_rechazo')
    
    # 1. Notificar al profesor por correo
    enviar_rechazo_profesor(usuario, motivo, request.host_url)
    
    # 2. Eliminar registros y archivos asociados (Opcional pero recomendado)
    # Aquí podrías borrar los archivos de static/uploads si lo deseas
    db.session.delete(profe)
    db.session.delete(usuario)
    db.session.commit()
    
    flash("Solicitud denegada y registros eliminados.", "warning")
    return redirect(url_for('admin_profesor_revision'))

# ==========================================================
# CORREOS PARA PROFESORES
# ==========================================================

# Mensaje automatico tras hacer el registro
# Mensaje automático tras hacer el registro (Estructura Original Mantenida)
def enviar_correo_recepcion_profesor(profesor, dominio):
    msg = MIMEMultipart('alternative')
    msg['From'] = f"Recursos Humanos UNGE <{CORREO_MATRICULAS_USER}>"
    # Usamos el correo personal que el aspirante acaba de registrar
    msg['To'] = profesor.correo_personal
    msg['Subject'] = f"Solicitud de Registro Recibida - Ref: {profesor.id}"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #1a237e; font-weight: bold;">Departamento de Recursos Humanos</p>
                </td>
            </tr>
            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #ff6f00;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: 300;">SOLICITUD EN REVISIÓN</h2>
                </td>
            </tr>
            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a Prof. <strong>{profesor.nombre_aspirante} {profesor.apellidos_aspirante}</strong>,</p>
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Le informamos que hemos recibido satisfactoriamente su solicitud de registro y la documentación adjunta (DIP y Credenciales). 
                        Actualmente, nuestro equipo está verificando la autenticidad de los datos.
                    </p>
                    <div style="background-color: #f8fafb; border-radius: 10px; padding: 25px; border-left: 5px solid #1a237e; margin-bottom: 30px;">
                        <table width="100%" style="font-size: 14px; color: #37474f;">
                            <tr><td><strong>ID DE SOLICITUD:</strong></td><td style="text-align: right;">#{profesor.id}</td></tr>
                            <tr><td><strong>DIP:</strong></td><td style="text-align: right;">{profesor.dip_aspirante}</td></tr>
                            <tr><td><strong>DEPARTAMENTO:</strong></td><td style="text-align: right;">{profesor.departamento}</td></tr>
                        </table>
                    </div>
                    <p style="font-size: 14px; color: #455a64; text-align: center;">
                        Una vez aprobada su solicitud, recibirá un segundo correo con su <strong>cuenta institucional</strong> y su <strong>código de activación</strong>.
                    </p>
                </td>
            </tr>
            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center;">
                    <p style="margin: 0; color: #78909c; font-size: 12px;"><strong>UNGE - Gestión de Profesorado</strong><br>Este es un mensaje automático.</p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))

    # BLOQUE DE ENVÍO CONFIGURADO PARA MAILHOG
    try:
        with smtplib.SMTP('127.0.0.1', 1025) as server:
            server.send_message(msg)
            print(f"DEBUG: Correo de recepción enviado a {profesor.correo_personal}")
    except Exception as e:
        print(f"DEBUG ERROR: No se pudo enviar el correo: {e}")


# Correo de activacion de la cuenta
# Correo de activacion de la cuenta (Estructura Original Mantenida)
def enviar_activacion_profesor(aspirante, codigo, correo_inst, dominio):
    msg = MIMEMultipart('alternative')
    msg['From'] = f"Soporte Técnico UNGE <{CORREO_MATRICULAS_USER}>"
    # IMPORTANTE: Se envía al correo personal para que pueda verlo y activar
    msg['To'] = aspirante.correo_personal
    msg['Subject'] = "¡ALTA CONFIRMADA! Active su Cuenta de Docente"

    # Ajustamos la URL para que coincida con nuestra nueva estructura limpia
    url_activacion = f"{dominio.rstrip('/')}/profesor/activar"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #28a745; font-weight: bold;">Validación de Identidad Docente</p>
                </td>
            </tr>
            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #1a237e;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: 300;">¡CUENTA VALIDADA!</h2>
                </td>
            </tr>
            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a <strong>{aspirante.nombre_aspirante}</strong>,</p>
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6;">Su identidad ha sido confirmada. Se ha generado su nueva identidad digital para el acceso a la plataforma académica:</p>
                    
                    <div style="background-color: #f8fafb; border-radius: 10px; padding: 25px; border-left: 5px solid #28a745; margin: 25px 0;">
                        <p style="margin: 0; font-size: 13px;"><strong>CORREO INSTITUCIONAL:</strong></p>
                        <p style="font-size: 18px; color: #1a237e; font-weight: bold; margin: 5px 0;">{correo_inst}</p>
                        <p style="margin: 15px 0 0 0; font-size: 13px;"><strong>CÓDIGO DE ACTIVACIÓN:</strong></p>
                        <p style="font-size: 22px; color: #ff6f00; font-weight: bold; letter-spacing: 4px; margin: 5px 0;">{codigo}</p>
                    </div>

                    <div style="text-align: center; margin: 40px 0;">
                        <a href="{url_activacion}" 
                           style="background-color: #1a237e; color: #ffffff; padding: 18px 35px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                            ACTIVAR MI CUENTA DOCENTE
                        </a>
                    </div>
                </td>
            </tr>
            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center;">
                    <p style="margin: 0; color: #78909c; font-size: 11px;">Este código es de un solo uso. Si no reconoce esta solicitud, contacte con soporte técnico.</p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))

    # BLOQUE DE ENVÍO SMTP (Mantenemos MailHog)
    try:
        with smtplib.SMTP('127.0.0.1', 1025) as server:
            server.send_message(msg)
            print(f"DEBUG: Correo de activación enviado a {aspirante.correo_personal}")
    except Exception as e:
        print(f"DEBUG ERROR: No se pudo enviar el correo de activación: {e}")


# REGISTRO RECHAZADO
def enviar_rechazo_profesor(usuario, motivo, dominio):
    msg = MIMEMultipart('alternative')
    msg['From'] = f"Recursos Humanos UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = usuario.correo
    msg['Subject'] = "IMPORTANTE: Solicitud de Registro Denegada"

    # Ajustamos el enlace para que el profesor pueda volver a intentarlo 
    # en la ruta correcta que configuramos antes
    url_reintento = f"{dominio.rstrip('/')}/profesor/registro"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <h1 style="margin: 0; font-size: 14px; color: #b71c1c; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #b71c1c; font-weight: bold;">Departamento de Recursos Humanos</p>
                </td>
            </tr>
            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #b71c1c;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 24px; font-weight: 300;">SOLICITUD RECHAZADA</h2>
                </td>
            </tr>
            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado/a <strong>{usuario.nombre}</strong>,</p>
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6;">
                        Tras revisar su documentación, lamentamos informarle que su solicitud de alta como docente ha sido <strong>RECHAZADA</strong> por el siguiente motivo:
                    </p>
                    
                    <div style="background-color: #fff5f5; border-radius: 10px; padding: 25px; border-left: 5px solid #b71c1c; margin: 25px 0; color: #b71c1c; font-weight: bold;">
                        {motivo}
                    </div>

                    <p style="font-size: 14px; color: #455a64;">
                        Para subsanar este error, deberá realizar un nuevo registro con la documentación correcta o ponerse en contacto con la secretaría académica.
                    </p>

                    <div style="text-align: center; margin: 40px 0;">
                        <a href="{url_reintento}" 
                           style="background-color: #2c3e50; color: #ffffff; padding: 18px 35px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                           INTENTAR REGISTRO DE NUEVO
                        </a>
                    </div>
                </td>
            </tr>
            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center;">
                    <p style="margin: 0; color: #78909c; font-size: 11px;">UNGE - Sistema de Gestión de Credenciales</p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))

    # BLOQUE DE ENVÍO SMTP
    try:
        with smtplib.SMTP('127.0.0.1', 1025) as server:
            server.send_message(msg)
            print(f"DEBUG: Correo de rechazo enviado a {usuario.correo}")
    except Exception as e:
        print(f"DEBUG ERROR: No se pudo enviar el correo de rechazo: {e}")




# ==========================================================
# SECCION PARA LA SECRETARIA
# ==========================================================

@app.route('/secretaria/panel')
@requiere_login
@requiere_rol(['secretaria'])
def panel_secretaria():
    # Gracias a tu relación: secretaria = db.relationship('Secretaria', backref='usuario', uselist=False)
    # podemos acceder directamente así:
    datos_secretaria = current_user.secretaria
    
    if not datos_secretaria:
        # Si por alguna razón no tiene perfil, lo redirigimos o manejamos el error
        flash("Perfil de secretaria no encontrado.", "warning")
        return redirect(url_for('inicio')) # o la página que prefieras
        
    return render_template('secretaria_panel.html', s=datos_secretaria)

# EDITAR PERFIL 
from werkzeug.security import generate_password_hash, check_password_hash

@app.route('/secretaria/editar-perfil', methods=['POST'])
@login_required
def directivo_editar_perfil():
    perfil = current_user.secretaria
    
    # --- 1. DATOS PROFESIONALES ---
    perfil.departamento = request.form.get('departamento').upper()
    perfil.titulo = request.form.get('titulo').upper()

    # --- 2. SEGURIDAD (CONTRASEÑA) ---
    pass_actual = request.form.get('password_actual')
    nueva_pass = request.form.get('nueva_password')
    confirm_pass = request.form.get('confirm_password')

    if nueva_pass:  # Si intentó cambiar la clave
        # A. Verificar si puso la clave actual correcta
        if not check_password_hash(current_user.password_hash, pass_actual):
            flash("La contraseña actual es incorrecta. No se realizaron cambios.", "danger")
            return redirect(url_for('panel_secretaria'))
        
        # B. Verificar si coinciden las nuevas
        if nueva_pass != confirm_pass:
            flash("Las nuevas contraseñas no coinciden entre sí.", "danger")
            return redirect(url_for('panel_secretaria'))
        
        # C. Verificar longitud
        if len(nueva_pass) < 6:
            flash("La nueva contraseña debe tener al menos 6 caracteres.", "warning")
            return redirect(url_for('panel_secretaria'))

        # Si todo bien, actualizamos el hash
        current_user.password_hash = generate_password_hash(nueva_pass)
        flash("Contraseña actualizada correctamente.", "success")

    # --- 3. FOTO ---
    # (Tu lógica de fotos aquí...)

    try:
        db.session.commit()
        flash("Perfil actualizado correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error crítico al guardar en la base de datos.", "danger")

    return redirect(url_for('panel_secretaria'))

    # 3. Foto de Perfil
    foto = request.files.get('archivo_foto')
    if foto and foto.filename != '':
        filename = secure_filename(f"sec_{current_user.id}_{foto.filename}")
        # Asegúrate de que esta carpeta exista en tu servidor
        ruta = os.path.join('static/uploads/perfiles', filename)
        foto.save(ruta)
        
        # Sincronizamos ambas tablas
        current_user.foto_perfil = filename
        perfil.foto_perfil = filename

    try:
        db.session.commit()
        flash("Perfil actualizado con éxito.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error al guardar los cambios.", "danger")

    return redirect(url_for('panel_secretaria'))


# FUNCION AUXILIAR FILTRO-CREACION
def obtener_datos_estudiantes_procesados(filtro_titu=None, filtro_asig=None):
    todos_los_profesores = Profesor.query.all()
    agrupado_estudiantes = {}
    titulaciones_set = set()
    asignaturas_set = set()

    base_dir = os.path.abspath(os.path.dirname(__file__))
    carpeta_entregas = os.path.join(base_dir, 'static', 'uploads', 'entregas')

    for profe in todos_los_profesores:
        if not profe.entregas_oficiales_json: continue
        try:
            data = json.loads(profe.entregas_oficiales_json)
            for entrega in data.get("entregas", []):
                if entrega.get("estado") == 'Recibido':
                    archivo_nombre = entrega.get('archivo', '')
                    ruta_excel = os.path.join(carpeta_entregas, archivo_nombre)
                    
                    if os.path.exists(ruta_excel):
                        df = pd.read_excel(ruta_excel)
                        # Normalizamos columnas: minúsculas, sin espacios y sin tildes para evitar errores
                        df.columns = [str(c).strip().lower().replace('é','e').replace('ó','o') for c in df.columns]
                        
                        if 'dip' in df.columns and 'nota' in df.columns:
                            titu = entrega.get('titulacion', 'Sin Titulación')
                            asig = entrega.get('asignatura', 'Sin Asignatura')
                            titulaciones_set.add(titu)
                            asignaturas_set.add(asig)

                            if filtro_titu and filtro_titu != titu: continue
                            if filtro_asig and filtro_asig != asig: continue

                            for _, fila in df.iterrows():
                                dip_str = str(fila['dip']).split('.')[0].strip()
                                try: 
                                    nota_val = float(fila.get('nota', 0))
                                except: 
                                    nota_val = 0

                                if dip_str not in agrupado_estudiantes:
                                    user_db = Usuario.query.filter_by(dip=dip_str).first()
                                    agrupado_estudiantes[dip_str] = {
                                        'nombre_completo': f"{user_db.apellidos}, {user_db.nombre}" if user_db else f"No en sistema ({dip_str})",
                                        'dip': dip_str,
                                        'titulacion': titu,
                                        'materias': []
                                    }
                                
                                # CAPTURA DE NUEVOS CAMPOS PARA EL ACTA OFICIAL
                                agrupado_estudiantes[dip_str]['materias'].append({
                                    'asignatura': asig,
                                    'nota': nota_val,
                                    'creditos': fila.get('creditos', '---'), # Captura columna 'creditos'
                                    'tipo': str(fila.get('tipo', 'OB')).upper(), # Captura 'OB' o 'OP'
                                    'convocatoria': str(fila.get('convocatoria', 'ORDINARIA')).upper(),
                                    'anio': entrega.get('curso', 'N/A'),
                                    'semestre': entrega.get('periodo', 'N/A'),
                                    'estado': "Aprobado" if nota_val >= 5 else "Suspenso"
                                })
        except Exception as e: 
            print(f"Error procesando actas: {e}")
    
    return agrupado_estudiantes, sorted(list(titulaciones_set)), sorted(list(asignaturas_set))

#1. PANEL PARA ELABORAR ACTAS + LOGICA
@app.route('/secretaria/actas/panel-elaboracion')
@login_required
def panel_elaboracion_actas():
    if current_user.rol not in ['admin', 'administrador', 'secretaria', 'profesor']:
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('inicio'))

    filtro_titu = request.args.get('titulacion')
    filtro_asig = request.args.get('asignatura')

    # Llamamos a la función auxiliar
    agrupado, titulaciones, asignaturas = obtener_datos_estudiantes_procesados(filtro_titu, filtro_asig)
    
    # --- LOGICA NUEVA PARA LA FOTO ---
    lista_final = []
    for dip, info in agrupado.items():
        # Buscamos al usuario en la DB para obtener su foto real
        estudiante_db = Usuario.query.filter_by(dip=dip).first()
        
        # Agregamos el campo foto_perfil al diccionario que va al HTML
        info['foto_perfil'] = estudiante_db.foto_perfil if estudiante_db else None
        lista_final.append(info)
    
    # Ordenamos la lista ya con las fotos incluidas
    lista_final = sorted(lista_final, key=lambda x: x['nombre_completo'])

    return render_template('panel_actas_secretaria.html', 
                           estudiantes_detectados=lista_final,
                           titulaciones=titulaciones,
                           asignaturas=asignaturas,
                           filtro_titu=filtro_titu,
                           filtro_asig=filtro_asig)


# 2. LOGICA QUE PROCESA LA ELABORACION OFICIAL DE ACTAS DE NOTAS 
@app.route('/secretaria/generar-acta/<dip>')
@login_required
def generar_acta_estudiante(dip):
    estudiante = Usuario.query.filter_by(dip=dip).first()
    if not estudiante:
        flash("Estudiante no encontrado.", "danger")
        return redirect(url_for('panel_elaboracion_actas'))

    agrupado, _, _ = obtener_datos_estudiantes_procesados() 
    datos_estudiante = agrupado.get(str(dip))

    if not datos_estudiante:
        flash("No hay notas procesadas.", "warning")
        return redirect(url_for('panel_elaboracion_actas'))

    # 1. Definimos los mapeos y el orden
    orden_logico = {'PRIMER': 1, 'SEGUNDO': 2, 'TERCERO': 3, 'CUARTO': 4, 'QUINTO': 5, 'SEXTO': 6}
    
    mapeo_niveles = {
        'PRIMER SEMESTRE': 'PRIMER', 'SEGUNDO SEMESTRE': 'PRIMER',
        'TERCER SEMESTRE': 'SEGUNDO', 'CUARTO SEMESTRE': 'SEGUNDO',
        'QUINTO SEMESTRE': 'TERCERO', 'SEXTO SEMESTRE': 'TERCERO',
        'SEPTIMO SEMESTRE': 'CUARTO', 'OCTAVO SEMESTRE': 'CUARTO',
        'NOVENO SEMESTRE': 'QUINTO', 'DECIMO SEMESTRE': 'QUINTO',
        'UNDECIMO SEMESTRE': 'SEXTO', 'DUODECIMO SEMESTRE': 'SEXTO'
    }

    historial_por_nivel = {}
    suma_global = 0.0
    conteo_global = 0
    
    # 2. Procesamos las materias
    for m in datos_estudiante['materias']:
        nombre_semestre = str(m.get('semestre') or 'SIN SEMESTRE').strip().upper()
        nivel_texto = mapeo_niveles.get(nombre_semestre, 'PRIMER')
        
        if nivel_texto not in historial_por_nivel:
            historial_por_nivel[nivel_texto] = {
                'anio_texto': str(m.get('anio') or 'N/A').strip(),
                'semestres': {}, 
                'suma_notas': 0.0,
                'conteo': 0
            }
        
        if nombre_semestre not in historial_por_nivel[nivel_texto]['semestres']:
            historial_por_nivel[nivel_texto]['semestres'][nombre_semestre] = []
        
        historial_por_nivel[nivel_texto]['semestres'][nombre_semestre].append(m)
        
        try:
            nota_valor = float(m.get('nota', 0))
            historial_por_nivel[nivel_texto]['suma_notas'] += nota_valor
            historial_por_nivel[nivel_texto]['conteo'] += 1
            # Actualizamos totales globales para la media
            suma_global += nota_valor
            conteo_global += 1
        except (ValueError, TypeError):
            pass

    # 3. CALCULAMOS LA MEDIA GENERAL (IMPORTANTE: Antes del return)
    media_general = suma_global / conteo_global if conteo_global > 0 else 0.0


    # Calculamos el año de inicio desde la fecha de creación del usuario
    # Si no tiene fecha, usamos el año actual como fallback
    anio_inicio = estudiante.fecha_creacion.year if estudiante.fecha_creacion else datetime.now().year
    anio_fin = anio_inicio + 1
    
    # Creamos el formato "2024-2025"
    expediente_dinamico = f"{anio_inicio}-{anio_fin}"

    # 4. ORDENAMOS EL HISTORIAL
    historial_ordenado = dict(sorted(
        historial_por_nivel.items(), 
        key=lambda x: orden_logico.get(x[0], 99)
    ))

    # 5. ENVIAMOS TODO AL HTML
    return render_template('formato_acta_oficial.html', 
                           e=estudiante,
                           usuario=estudiante, # Para que capte usuario.natural_de, etc.
                           historial=historial_ordenado,
                           media_general=media_general, # Ahora sí está definida
                           anio_ingreso=anio_inicio,
                           anio_expediente=expediente_dinamico, # <--- Enviamos el dinámico
                           fecha=datetime.now())


# SIBIDA DE ARCHIVOS

@app.route('/almacenamiento/archivo-digital')
@login_required
@requiere_rol(['secretaria', 'directivo'])
def gestionar_archivo():
    # LÓGICA DE PRIVACIDAD: 
    # Vemos carpetas que son públicas O las que yo mismo he creado (aunque sean privadas)
    carpetas_visibles = Carpeta.query.filter(
        or_(
            Carpeta.es_publica == True,
            Carpeta.creador_id == current_user.id
        )
    ).order_by(Carpeta.fecha_creacion.desc()).all()
    
    return render_template('subir_documento_secretaria.html', carpetas=carpetas_visibles)


# Manejo de archivos y guardado
UPLOAD_FOLDER_DOCS = 'static/uploads/archivo_digital'
os.makedirs(UPLOAD_FOLDER_DOCS, exist_ok=True)

# Extensiones permitidas (solo documentos)
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/almacenamiento/archivo/crear-carpeta', methods=['POST'])
@login_required
@requiere_rol(['secretaria', 'directivo'])
def crear_carpeta():
    nombre = request.form.get('nombre_carpeta')
    # Si el checkbox 'es_publica' viene en el form, será True, si no, False
    privacidad_radio = request.form.get('privacidad') 
    publica = True if privacidad_radio == 'publica' else False

    if nombre:
        nueva_carpeta = Carpeta(
            nombre=nombre.strip().upper(),
            creador_id=current_user.id,  # Guardamos quién la creó
            es_publica=publica           # Definimos si otros la verán
        )
        try:
            db.session.add(nueva_carpeta)
            db.session.commit()
            flash(f'Carpeta "{nombre}" creada como {"Pública" if publica else "Privada"}.', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error al crear la carpeta.', 'danger')
            
    return redirect(url_for('gestionar_archivo'))

@app.route('/almacenamiento/archivo/subir/<int:carpeta_id>', methods=['POST'])
@login_required
@requiere_rol(['secretaria', 'directivo'])
def subir_a_carpeta(carpeta_id):
    # Verificación de seguridad: ¿Tiene el usuario permiso para subir a esta carpeta?
    carpeta = Carpeta.query.get_or_404(carpeta_id)
    if not carpeta.es_publica and carpeta.creador_id != current_user.id:
        flash('No tienes permiso para subir archivos a esta carpeta privada.', 'danger')
        return redirect(url_for('gestionar_archivo'))

    if 'archivo' not in request.files:
        flash('No se seleccionó ningún archivo.', 'warning')
        return redirect(url_for('gestionar_archivo'))
    
    archivo = request.files['archivo']
    if archivo and allowed_file(archivo.filename):
        original_name = secure_filename(archivo.filename)
        unique_name = f"{uuid.uuid4().hex[:8]}_{original_name}"
        ruta = os.path.join(UPLOAD_FOLDER_DOCS, unique_name)
        
        try:
            archivo.save(ruta)
            nuevo_doc = DocumentoArchivo(
                nombre_archivo=unique_name, 
                carpeta_id=carpeta_id,
                # Opcional: podrías guardar también quién subió el archivo
                # subido_por=current_user.id 
            )
            db.session.add(nuevo_doc)
            db.session.commit()
            flash('Documento archivado correctamente.', 'success')
        except Exception as e:
            db.session.rollback()
            if os.path.exists(ruta):
                os.remove(ruta)
            flash(f'Error de sistema al registrar el archivo.', 'danger')
    else:
        flash('Tipo de archivo no permitido.', 'danger')
    
    return redirect(url_for('gestionar_archivo'))


# Eliminar la carpeta (usuario_id)

@app.route('/almacenamiento/carpeta/eliminar/<int:id>', methods=['POST'])
@login_required
@requiere_rol(['secretaria', 'directivo'])
def eliminar_carpeta(id):
    # 1. Buscar la carpeta o devolver 404 si no existe
    carpeta = Carpeta.query.get_or_404(id)

    # 2. SEGURIDAD: Solo el creador original puede borrarla
    if carpeta.creador_id != current_user.id:
        flash("No tienes permiso para eliminar esta carpeta, solo su creador puede hacerlo.", "danger")
        return redirect(url_for('gestionar_archivo'))

    try:
        # 3. Borrar los archivos físicos del disco antes de borrar de la DB
        for doc in carpeta.documentos:
            ruta_archivo = os.path.join(UPLOAD_FOLDER_DOCS, doc.nombre_archivo)
            if os.path.exists(ruta_archivo):
                os.remove(ruta_archivo)
            # El registro del documento se borrará automáticamente si usas cascade en tu modelo
            # Si no, bórralo manualmente: db.session.delete(doc)

        # 4. Borrar la carpeta de la base de datos
        db.session.delete(carpeta)
        db.session.commit()
        
        flash(f'Carpeta "{carpeta.nombre}" y todos sus archivos han sido eliminados.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al intentar eliminar la carpeta: {str(e)}', 'danger')
        print(f"Error en eliminación: {e}")

    return redirect(url_for('gestionar_archivo'))


# ==========================================================
# PANEL PARA DIRECTIVOS
# ==========================================================

# PANEL ADMINISTRATIVO
@app.route('/decano/panel')
@requiere_login
@requiere_rol('directivo') # Aquí usamos el rol directivo que definiste antes
def directivo_panel():
    # Datos del decano para mostrar en las credenciales
    datos_decano = Profesor.query.filter_by(usuario_id=current_user.id).first()

    # Contador de usuarios
    usuarios_total = Usuario.query.count()
    return render_template('directivo_panel.html', d=datos_decano, usuarios_count=usuarios_total)

# VISTA DEL DIRECTIVO PANEL ADMINISTRAVIO
@app.route('/panel/directivo/<int:directivo_id>')
def ver_perfil_directivo(directivo_id):
    # Buscamos el directivo por su ID
    # .get_or_404() hace que si no existe, muestre una página de error limpia
    directivo = Directivo.query.get_or_404(directivo_id)
    
    # Renderizamos el perfil público (asegúrate de que el nombre del archivo coincida)
    return render_template('perfil_publico_decano.html', d=directivo)

# PERFIL DE LOS DIRECTIVOS.
@app.route('/perfil/directivo/<int:user_id>')
@login_required
def perfil_directivo_publico(user_id):
    print(f"DEBUG: Intentando acceder al perfil directivo del ID: {user_id}")
    u = Usuario.query.get_or_404(user_id)
    return render_template('directivo_perfil.html', usuario=u)


# DOCUEMNTOS RECIBIDOS
@app.route('/directivo/reportes-expedientes')
@login_required
@requiere_rol(['directivo'])
def ver_reportes_expedientes():
    # Obtener parámetros de búsqueda
    search_name = request.args.get('remitente', '')
    search_cat = request.args.get('categoria', '')

    query = DocumentoRecibido.query.filter_by(destinatario_id=current_user.id)

    if search_name:
        # Filtramos por el nombre del remitente (asumiendo relación con Usuario)
        query = query.join(Usuario, DocumentoRecibido.remitente_id == Usuario.id)\
                     .filter(Usuario.nombre.ilike(f'%{search_name}%'))
    
    if search_cat:
        query = query.filter(DocumentoRecibido.categoria == search_cat)

    reportes = query.order_by(DocumentoRecibido.fecha_envio.desc()).all()
    
    # Necesitamos las carpetas del directivo para el botón "Archivar"
    carpetas = Carpeta.query.filter_by(creador_id=current_user.id).all()
    
    return render_template('directivo_reportes_recibidos.html', reportes=reportes, carpetas=carpetas)


# Panel que poseen terceros para hacer envio
@app.route('/enviar-reporte', methods=['GET', 'POST'])
@login_required
def enviar_reporte():
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        categoria = request.form.get('categoria')
        destinatario_id = request.form.get('destinatario_id')
        archivo = request.files.get('archivo')

        if archivo and allowed_file(archivo.filename):
            # 1. Intentamos guardar el archivo físico
            ruta_relativa = guardar_archivo(archivo, 'recibidos')
            
            if ruta_relativa:
                nombre_final = ruta_relativa.split('/')[-1]

                # 2. Creamos el objeto para la DB
                nuevo_reporte = DocumentoRecibido(
                    titulo=titulo.strip().upper() if titulo else "SIN TITULO",
                    archivo_nombre=nombre_final,
                    remitente_id=current_user.id,
                    destinatario_id=destinatario_id,
                    categoria=categoria,
                    fecha_envio=datetime.utcnow()
                )
                
                try:
                    db.session.add(nuevo_reporte)
                    db.session.commit()
                    flash("Reporte enviado correctamente.", "success")
                    return redirect(url_for('perfil_directivo_publico', user_id=destinatario_id))
                
                except Exception as e:
                    db.session.rollback()
                    # Si falla la DB, borramos el archivo físico para no dejar basura
                    ruta_fisica = os.path.join(STATIC_DIR, ruta_relativa)
                    if os.path.exists(ruta_fisica):
                        os.remove(ruta_fisica)
                    
                    # IMPRIMIR EL ERROR REAL EN CONSOLA
                    print("---------- ERROR DE BASE DE DATOS ----------")
                    print(str(e)) 
                    print("--------------------------------------------")
                    
                    flash(f"Atención: Error al registrar en la base de datos. Verifique los campos.", "danger")
            else:
                flash("Error al guardar el archivo físico.", "danger")
        else:
            flash("Archivo no permitido.", "warning")

    directivos = Usuario.query.filter_by(rol='directivo').all()
    return render_template('enviar_reporte_directivo.html', directivos=directivos)

# Valodar los reportes.
@app.route('/validar-reporte/<int:id>/<estado>', methods=['POST'])
@login_required
def validar_reporte(id, estado):
    reporte = DocumentoRecibido.query.get_or_404(id)
    
    # --- NUEVO: GUARDAR EL ESTADO EN EL REPORTE ---
    # Esto permite que el HTML sepa qué botón bloquear
    reporte.estado = estado 
    # ----------------------------------------------

    # Configuramos el contenido según el estado
    if estado == 'validado':
        texto = f"✅ Tu reporte '{reporte.titulo}' ha sido VALIDADO correctamente."
        categoria_notif = 'success'
    else:
        texto = f"❌ Tu reporte '{reporte.titulo}' ha sido RECHAZADO. Por favor, revisa los requisitos y vuelve a enviarlo."
        categoria_notif = 'rechazado'

    try:
        # 1. Creamos la Notificación
        nueva_notificacion = Notificacion(
            usuario_id=reporte.remitente_id,
            tipo=categoria_notif,
            mensaje=texto,
            item_id=reporte.id,
            leida=False
        )

        # 2. Creamos el Mensaje
        nuevo_mensaje = Mensaje(
            emisor_id=current_user.id,
            receptor_id=reporte.remitente_id,
            contenido=texto,
            leido=False
        )

        db.session.add(nueva_notificacion)
        db.session.add(nuevo_mensaje)
        db.session.commit() # Aquí se guarda tanto el estado del reporte como la notif.
        
        flash(f"Documento {estado} y usuario notificado.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"Error al notificar: {e}")
        flash("Error al procesar la validación.", "danger")

    return redirect(request.referrer)

# Eliminar un reporte
@app.route('/eliminar-reporte/<int:id>', methods=['POST'])
@login_required
def eliminar_reporte(id):
    reporte = DocumentoRecibido.query.get_or_404(id)
    
    # 1. Borrar archivo físico
    # Usamos la lógica de tu STATIC_DIR para localizarlo exactamente
    ruta_archivo = os.path.join(STATIC_DIR, 'uploads', 'recibidos', reporte.archivo_nombre)
    
    try:
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
        
        # 2. Borrar de la DB
        db.session.delete(reporte)
        db.session.commit()
        flash("Reporte eliminado definitivamente.", "info")
    except Exception as e:
        db.session.rollback()
        flash("Error al eliminar el archivo.", "danger")
        
    return redirect(request.referrer)

# EDITAR PERFIL DEL DIRECTIVO
# Cambiar contrasena
@app.route('/editar-perfil-directivo', methods=['POST'])
@login_required
def editar_perfil_directivo():
    # 1. Obtener el perfil directivo vinculado
    perfil = current_user.directivo
    if not perfil:
        perfil = Directivo(usuario_id=current_user.id)
        db.session.add(perfil)

    # 2. Actualizar datos de contacto (Tabla Usuario)
    current_user.correo = request.form.get('correo')
    current_user.telefono = request.form.get('telefono')
    
    # 3. Actualizar datos de cargo (Tabla Directivo)
    perfil.cargo = request.form.get('cargo')
    perfil.ubicacion = request.form.get('ubicacion')
    # Añadimos seccion si también quieres editarla (puedes añadir el input al HTML)
    perfil.seccion = request.form.get('seccion')

    # 4. GESTIÓN DE SEGURIDAD (Cambio de contraseña)
    old_pass = request.form.get('old_password')
    new_pass = request.form.get('new_password')
    confirm_pass = request.form.get('confirm_password')

    if old_pass: # Si el usuario intentó cambiar la clave
        if not current_user.check_password(old_pass):
            flash("La contraseña actual no es correcta.", "danger")
            return redirect(request.referrer)
        
        if new_pass != confirm_pass:
            flash("Las nuevas contraseñas no coinciden.", "danger")
            return redirect(request.referrer)
        
        if len(new_pass) < 8:
            flash("La nueva contraseña debe tener al menos 8 caracteres.", "warning")
            return redirect(request.referrer)
        
        current_user.set_password(new_pass)
        flash("Contraseña actualizada con éxito.", "success")

    # 5. Gestión de Foto
    file = request.files.get('foto')
    if file and file.filename != '':
        filename = secure_filename(f"perfil_{current_user.id}_{file.filename}")
        file.save(os.path.join('static/uploads/perfiles', filename))
        current_user.foto_perfil = filename

    try:
        db.session.commit()
        flash("Perfil actualizado correctamente.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al guardar los cambios: {e}", "danger")

    return redirect(request.referrer)

# ==========================================================
# ADMINSTRACION GENERAL DE LA PLATAFORMA
# ==========================================================

# PANEL DEL ADMIN
@app.route('/panel_admin')
@login_required
@requiere_rol("admin")
def panel_admin():
    # Solo permitimos el acceso si el rol es 'admin' o 'administrador'
    if current_user.rol not in ['admin', 'administrador']:
        return "Acceso denegado", 403

    # Recolección de datos reales para el Dashboard
    total_est = Usuario.query.filter_by(rol='estudiante').count()
    total_prof = Usuario.query.filter_by(rol='profesor').count()
    total_direc = Usuario.query.filter_by(rol='directivo').count()
    total_admin = Usuario.query.filter_by(rol='admin').count()
    pendientes = 12  # Esto vendrá de tu tabla de inscripciones luego
    buzon = 5       # Esto vendrá de tu tabla de mensajes/dudas
    
    # Obtenemos los últimos 5 usuarios registrados para la tabla
    ultimos = Usuario.query.order_by(Usuario.fecha_creacion.desc()).limit(5).all()

    return render_template('admin_panel.html', 
                           total_estudiantes=total_est,
                           total_profesores=total_prof,
                           total_directivos=total_direc,
                           total_administradores=total_admin,
                           solicitudes_pendientes=pendientes,
                           dudas_buzon=buzon,
                           ultimos_usuarios=ultimos)


# SU RUTA PARA HACER LOGIN

@app.route('/admin/login', methods=['GET', 'POST'])
def login_admin():

    # Contamos cada tipo de usuario
    total_est = Usuario.query.filter_by(rol='estudiante').count()
    total_prof = Usuario.query.filter_by(rol='profesor').count()
    total_dir = Usuario.query.filter_by(rol='directivo').count()
    total_adm = Usuario.query.filter_by(rol='admin').count()

    # Otros datos (puedes dejarlos en 0 por ahora o contarlos si tienes las tablas)
    pendientes = 0 
    buzon = 0
    ultimos = Usuario.query.order_by(Usuario.id.desc()).limit(5).all()
    # Si ya está logueado como admin, enviarlo al panel
    if current_user.is_authenticated and current_user.rol == 'admin':
        return redirect(url_for('panel_admin'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        # Buscamos por correo o por un campo 'username' si lo tienes
        user = Usuario.query.filter_by(correo_institucional=username).first()

        if user and check_password_hash(user.password_hash, password):
            if user.rol == 'admin':
                login_user(user)
                return redirect(url_for('panel_admin'))
            else:
                flash('Acceso denegado: Esta terminal es solo para administradores.', 'danger')
        else:
            flash('Credenciales de administrador incorrectas.', 'danger')

    return render_template('login_admin.html')




# ==========================================================
# DIRECTORIO DE PROFESORES
# ==========================================================

# LISTAR PROFESORES
@app.route('/directorio/profesores')
def directorio_profesores():
    # 1. Consulta unificada trayendo al Usuario y sus posibles perfiles adicionales
    docentes_raw = db.session.query(Usuario, Profesor, Directivo).outerjoin(
        Profesor, Usuario.id == Profesor.usuario_id
    ).outerjoin(
        Directivo, Usuario.id == Directivo.usuario_id
    ).filter(
        Usuario.rol.in_(['profesor', 'directivo', 'admin'])
    ).all()
    
    # 2. Extraer listas únicas para los filtros dinámicos del HTML
    # Filtramos valores None o vacíos para que el select se vea limpio
    departamentos_existentes = sorted(list(set([p.departamento for u, p, d in docentes_raw if p and p.departamento])))
    cargos_existentes = sorted(list(set([d.cargo for u, p, d in docentes_raw if d and d.cargo])))

    directorio = []
    
    for usuario, profe, directivo in docentes_raw:
        # LÓGICA DE FOTO:
        # Prioridad Directivo: tabla Usuario | Prioridad Profesor: tabla Profesor
        foto_final = None
        if usuario.rol == 'directivo':
            foto_final = usuario.foto_perfil
        elif usuario.rol == 'profesor' and profe:
            foto_final = profe.archivo_foto
        
        # Fallback general (si no hay foto en tabla Profesor, busca en Usuario)
        if not foto_final:
            foto_final = usuario.foto_perfil

        # Definir el subtítulo para que el filtro de JS tenga una cadena con qué comparar
        if usuario.rol == 'directivo' and directivo:
            subtitulo = directivo.cargo
        elif profe:
            subtitulo = profe.departamento
        else:
            subtitulo = usuario.rol.capitalize()

        directorio.append({
            'usuario': usuario,
            'profe': profe,
            'directivo': directivo,
            'foto_display': foto_final,
            'subtitulo': subtitulo
        })
    
    # 3. Enviamos el directorio y las listas de filtros al template
    return render_template(
        'directorio_profesores.html', 
        directorio=directorio, 
        departamentos=departamentos_existentes,
        cargos=cargos_existentes
    )

# ==========================================================
# DIRECTORIO DE ESTUDIANTES
# ==========================================================
@app.route('/directorio/estudiantes')
def directorio_estudiantes():
    # Consultamos solo la tabla Usuario filtrando por el rol 'estudiante'
    estudiantes = Usuario.query.filter_by(rol='estudiante').all()
    
    # Extraer carreras únicas directamente de la columna carrera en usuarios
    carreras_existentes = sorted(list(set([u.carrera for u in estudiantes if u.carrera])))

    directorio = []
    for u in estudiantes:
        directorio.append({
            'usuario': u,
            'foto_display': u.foto_perfil,
            # Usamos los campos de la tabla usuarios para el subtítulo
            'subtitulo': f"{u.carrera or 'Estudiante'} - {u.curso or ''}"
        })
    
    return render_template(
        'directorio_estudiantes.html', 
        directorio=directorio, 
        carreras=carreras_existentes
    )


# ==========================================================
# BLOG DE NOTAS PARA ESTUDIANTES
# ==========================================================

@app.route('/notas')
@app.route('/notas/<int:user_id>')
@login_required
def ver_notas(user_id=None):
    # Buscamos al dueño de los datos
    usuario_perfil = Usuario.query.get_or_404(user_id)

    # 1. CAPTURA EL ID (Ya sea de la URL /40 o del parámetro ?id=)
    if user_id is None:
        # Intentamos obtenerlo de ?id=40
        query_id = request.args.get('id', type=int)
        # Si no hay ?id, usamos el del usuario actual
        target_user_id = query_id if query_id else current_user.id
    else:
        target_user_id = user_id
    
    # 2. Permiso: Solo el dueño edita
    puedo_editar = (target_user_id == current_user.id)
    
    # 3. Obtener el nombre del dueño del perfil
    usuario_perfil = Usuario.query.get_or_404(target_user_id)

    # 4. INICIALIZAR LA VARIABLE
    datos_completos = []

    # 5. Filtrar notas por el ID del perfil que visitamos
    # AQUÍ ES DONDE SE LOGRA LA PRIVACIDAD
    notas_usuario = Nota.query.filter_by(usuario_id=target_user_id).all()
    
    # Obtenemos IDs de asignaturas únicas para este usuario
    asignaturas_ids = list(set([n.asignatura_id for n in notas_usuario]))
    
    for a_id in asignaturas_ids:
        asig = Asignatura.query.get(a_id)
        if not asig: continue
        
        notas_asig = [n for n in notas_usuario if n.asignatura_id == a_id]
        
        cps, sms, evs = [""]*10, [""]*10, [""]*10
        reaccion_actual = ""
        semestre_actual = "1" # Por defecto

        for n in notas_asig:
            if n.tipo == 'Práctica': cps[n.posicion-1] = n.contenido
            elif n.tipo == 'Seminario': sms[n.posicion-1] = n.contenido
            elif n.tipo == 'Evaluación': evs[n.posicion-1] = n.contenido
            if n.reaccion: reaccion_actual = n.reaccion
            if n.semestre: semestre_actual = n.semestre

        datos_completos.append({
            'asignatura_nombre': asig.nombre,
            'creditos': asig.creditos,
            'semestre': semestre_actual, 
            'lista_notas': cps + sms + evs,
            'reaccion': reaccion_actual
        })
    
    return render_template('notas.html', 
                           tabla_datos=datos_completos, 
                           puedo_editar=puedo_editar, 
                           nombre_perfil=usuario_perfil.nombre,
                           usuario=usuario_perfil)


# GUARDAR NOTAS INSERTADAS
@app.route('/guardar_matriz', methods=['POST'])
@login_required
def guardar_matriz():
    datos = request.get_json()
    try:
        for item in datos:
            nombre_asig = item['asignatura'].strip()
            if not nombre_asig:
                continue

            # 1. Buscar o Crear la asignatura vinculada a ESTE usuario
            # Agregamos usuario_id al filtro para que 'Matemáticas' del ID 40 
            # sea diferente a 'Matemáticas' del ID 39.
            asig = Asignatura.query.filter_by(
                nombre=nombre_asig, 
                usuario_id=current_user.id
            ).first()

            if not asig:
                asig = Asignatura(
                    nombre=nombre_asig, 
                    creditos=int(item.get('creditos') or 0),
                    usuario_id=current_user.id  # <--- Dueño de la asignatura
                )
                db.session.add(asig)
                db.session.flush() # Para obtener el asig.id de inmediato
            else:
                # Si ya existe, actualizamos los créditos por si cambiaron
                asig.creditos = int(item.get('creditos') or 0)

            # 2. BORRADO SEGURO: Borrar solo mis notas de esta materia
            # Esto garantiza que si el ID 39 guarda, NO toque nada del ID 40
            Nota.query.filter_by(
                asignatura_id=asig.id, 
                usuario_id=current_user.id
            ).delete()

            # 3. Guardar nuevas notas con MI ID de sesión
            for i, valor in enumerate(item['notas']):
                # Guardamos si hay texto o si hay una reacción seleccionada
                if (valor and valor.strip() != "") or item.get('reaccion'):
                    # Determinamos tipo según la posición en el array de 30
                    if i < 10:
                        tipo_n = 'Práctica'
                    elif i < 20:
                        tipo_n = 'Seminario'
                    else:
                        tipo_n = 'Evaluación'
                    
                    pos_n = (i % 10) + 1
                    
                    nueva = Nota(
                        usuario_id=current_user.id, # <--- El candado de seguridad
                        asignatura_id=asig.id,
                        tipo=tipo_n,
                        posicion=pos_n,
                        contenido=valor.strip() if valor else "",
                        reaccion=item.get('reaccion'),
                        semestre=item.get('semestre') # Guardamos el semestre también
                    )
                    db.session.add(nueva)
        
        db.session.commit()
        return jsonify({"status": "success", "message": "Datos guardados en tu perfil"}), 200

    except Exception as e:
        db.session.rollback()
        print(f"Error al guardar: {str(e)}") # Para que lo veas en la consola
        return jsonify({"status": "error", "message": str(e)}), 500




# ==========================================================
# FOROS Y DEBATES INTERACTIVOS
# ==========================================================
# RUTA PARA CREAR FOROR
@app.route('/foro')
@login_required
def ver_foro():
    debates = Debate.query.order_by(Debate.fecha_creacion.desc()).all()
    form = DebateForm() # Para el modal de creación
    return render_template('ver_foro.html', debates=debates, form=form)

# CREACION DE DEBATES
@app.route('/crear_debate', methods=['GET', 'POST'])
@login_required 
def crear_debate():
    form = DebateForm()
    if form.validate_on_submit():
        archivo_nombre = None
        tipo_archivo = None

        # Procesar archivo si existe
        if form.archivo.data:
            f = form.archivo.data
            archivo_nombre = secure_filename(f"{datetime.now().timestamp()}_{f.filename}")
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], archivo_nombre))
            
            ext = archivo_nombre.rsplit('.', 1)[1].lower()
            if ext in ['mp4', 'mov', 'avi']: tipo_archivo = 'video'
            elif ext in ['jpg', 'png', 'jpeg', 'gif']: tipo_archivo = 'imagen'
            else: tipo_archivo = 'documento'

        # 1. Crear el debate
        nuevo_debate = Debate(
            titulo=form.titulo.data,
            contenido=form.contenido.data,
            archivo=archivo_nombre,
            tipo_archivo=tipo_archivo,
            autor_id=current_user.id
        )
        db.session.add(nuevo_debate)
        
        # 2. Sumar +1 al contador de debates del perfil del usuario actual
        current_user.debate = (current_user.debate or 0) + 1
        
        # Flush para que el debate tenga ID antes de crear las notificaciones
        db.session.flush() 

        # 3. LÓGICA DE NOTIFICACIÓN PARA LA PRUEBA
        # Buscamos a todos los usuarios que NO sean el autor actual y NO sean admin
        usuarios_comunidad = Usuario.query.filter(
            Usuario.id != current_user.id, 
            Usuario.rol != 'admin'
        ).all()

        for usuario in usuarios_comunidad:
            notificacion = Notificacion(
                usuario_id=usuario.id,
                tipo='debate',
                item_id=nuevo_debate.id,
                mensaje=f"🔔 {current_user.nombre} publicó un nuevo debate: {nuevo_debate.titulo}",
                leida=False, # Esto activa el círculo rojo en la barra del otro usuario
                fecha_creacion=datetime.utcnow()
            )
            db.session.add(notificacion)

        db.session.commit()
        flash('Debate publicado. ¡Tus compañeros han sido notificados!', 'success')
        return redirect(url_for('ver_foro'))

    return render_template('ver_foro.html', form=form)

# HACER COMENTARIOS DURANTE LOS FOROS
@app.route('/debate/comentar/<int:debate_id>', methods=['POST'])
@login_required
def comentar(debate_id):
    contenido = request.form.get('contenido')
    if contenido and contenido.strip():
        nuevo_comentario = Comentario(
            debate_id=debate_id,
            autor_id=current_user.id,
            contenido=contenido
        )
        db.session.add(nuevo_comentario)
        db.session.commit()
    return redirect(url_for('ver_foro'))


# ==========================================================
# NOTIFICACIONES GENERALES
# ==========================================================

@app.route('/notificacion/ir/<int:notif_id>')
@login_required
def ir_a_notificacion(notif_id):
    notif = Notificacion.query.get_or_404(notif_id)
    
    # Seguridad: verificar que la notificación pertenece al usuario actual
    if notif.usuario_id != current_user.id:
        abort(403)

    # 1. Marcar como leída (esto resta 1 al contador visual)
    notif.leida = True
    
    # 2. Lógica de Autolimpieza: Borrar notificaciones leídas de más de 30 días
    limite = datetime.utcnow() - timedelta(days=30)
    Notificacion.query.filter(
        Notificacion.usuario_id == current_user.id,
        Notificacion.leida == True,
        Notificacion.fecha_creacion < limite
    ).delete()
    
    db.session.commit()

    # 3. Redirección dinámica según el tipo
    if notif.tipo == 'debate':
        return redirect(url_for('ver_foro', _anchor=f'debate-{notif.item_id}')) # O a la vista específica si tienes ID
    elif notif.tipo == 'evento' and notif.item_id:
        return redirect(url_for('ver_eventos', evento_id=notif.item_id))
    elif notif.tipo == 'noticia' and notif.item_id:
        return redirect(url_for('noticia_completa', noticia_id=notif.item_id))
    elif notif.tipo == 'mensaje' and notif.item_id:
        return redirect(url_for('ver_mensaje', id=notif.item_id))
    elif notif.tipo == 'anuncio_directivo':
        return redirect(url_for('ver_mensaje', id=notif.item_id))
    elif notif.tipo == 'success':
        return redirect(url_for('ver_mensaje', id=notif.item_id))
    elif notif.tipo == 'rechazado':
        return redirect(url_for('ver_mensaje', id=notif.item_id))
    
    return redirect(url_for('inicio'))



# ==========================================================
# MI FACULTAD, INTERACCIONES UNIVERSITARIAS
# ==========================================================
# RUTA QUE MUESTRA LA PAGINA
@app.route('/facultad')
@login_required
def facultad():
    search = request.args.get('search', '')
    # Usamos joinedload para traer la relación 'profesor' de una vez y evitar errores en el HTML
    query = Usuario.query.options(joinedload(Usuario.profesor)).filter(Usuario.rol != 'admin')
    if search:
        # Filtramos por nombre, carrera o dip, excluyendo admin
        usuarios_lista = Usuario.query.filter(
            Usuario.rol != 'admin',
            (Usuario.nombre.ilike(f'%{search}%')) | 
            (Usuario.carrera.ilike(f'%{search}%')) |
            (Usuario.dip.ilike(f'%{search}%'))
        ).all()
    else:
        usuarios_lista = Usuario.query.filter(Usuario.rol != 'admin').limit(12).all()
    
    return render_template('mi_facultad.html', usuarios_lista=usuarios_lista)



# ==========================================================
# CREACION DE ANUNCIOS. DECANOS
# ==========================================================
@app.route('/decano/publicar-anuncio', methods=['GET', 'POST'])
@login_required
def publicar_anuncio_directivo():
    perfil = Directivo.query.filter_by(usuario_id=current_user.id).first()
    if not perfil:
        abort(403)

    if request.method == 'POST':
        titulo = request.form.get('titulo')
        contenido = request.form.get('contenido')
        archivo = request.files.get('documento')

        nombre_archivo = None
        if archivo and archivo.filename != '':
            nombre_archivo = secure_filename(f"anuncio_{datetime.now().timestamp()}_{archivo.filename}")
            archivo.save(os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo))

        try:
            # Filtrar destinatarios (todos menos el decano actual)
            destinatarios = Usuario.query.filter(Usuario.id != current_user.id).all()
            total_destinatarios = len(destinatarios)

            # 1. Guardar en historial de anuncios
            nuevo_anuncio = AnuncioDirectivo(
                directivo_id=perfil.id,
                titulo=titulo,
                contenido=contenido,
                archivo_adjunto=nombre_archivo,
                alcance=total_destinatarios
            )
            db.session.add(nuevo_anuncio)
            db.session.flush()

            # 2. Envío masivo como DM y Notificación
            for u in destinatarios:
                nuevo_dm = Mensaje(
                    emisor_id=current_user.id,
                    receptor_id=u.id,
                    contenido=f"📢 COMUNICADO OFICIAL: {titulo}\n\n{contenido}",
                    archivo_adjunto=nombre_archivo,
                    enviado=True, recibido=True, leido=False
                )
                db.session.add(nuevo_dm)
                db.session.flush()

                notif = Notificacion(
                    usuario_id=u.id,
                    tipo='mensaje',
                    mensaje=f"Nuevo comunicado del {perfil.cargo}",
                    item_id=nuevo_dm.id,
                    leida=False
                )
                db.session.add(notif)
            
            db.session.commit()
            flash(f"Comunicado enviado a {total_destinatarios} usuarios.", "success")
            return redirect(url_for('publicar_anuncio_directivo'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    anuncios = AnuncioDirectivo.query.filter_by(directivo_id=perfil.id).order_by(AnuncioDirectivo.fecha_creacion.desc()).all()
    return render_template('decano_crear_anuncio.html', anuncios=anuncios)




# ==========================================================
# CONTROL DISCIPLINARIO DE CUENTAS ESTUDIANTILES
# ==========================================================
# Funcion para enviar correo tras suspender la cuenta
def enviar_correo_disciplinario(estudiante, accion, motivo, fecha_expiracion=None, dominio="http://localhost:5000"):
    """
    Envía una notificación disciplinaria con el diseño Premium de la UNGE.
    """
    email_destino = estudiante.correo
    nombre_alumno = f"{estudiante.nombre}{estudiante.apellidos}"  
    
    
    # Configuración dinámica según la acción
    if accion == 'activar':
        color_principal = "#1b5e20"  # Verde UNGE
        color_fondo_caja = "#f1f8e9"
        titulo_banner = "ACCESO REESTABLECIDO"
        saludo = "¡Buenas noticias!"
        mensaje_estado = "Nos complace informarle que su acceso a la plataforma académica ha sido <strong>REESTABLECIDO</strong> satisfactoriamente."
        
        if fecha_expiracion:
            detalle = f"Vigencia del acceso hasta: <strong>{fecha_expiracion.strftime('%d/%m/%Y')}</strong>"
        else:
            detalle = "Su cuenta cuenta ahora con <strong>Acceso Indefinido</strong>."
            
        btn_texto = "ENTRAR A MI CUENTA"
        url_btn = f"{dominio}/login"
    else:
        color_principal = "#b71c1c"  # Rojo Disciplinario
        color_fondo_caja = "#ffebee"
        titulo_banner = "CUENTA SUSPENDIDA"
        saludo = "Aviso Importante"
        mensaje_estado = "Le informamos que su cuenta de estudiante ha sido <strong>DESACTIVADA</strong> temporalmente por el departamento administrativo."
        detalle = f"<strong>Motivo de la medida:</strong><br>{motivo}"
        btn_texto = "CONTACTAR SOPORTE"
        url_btn = "mailto:soporte@unge.gq"

    msg = MIMEMultipart('alternative')
    msg['From'] = f"Control Disciplinario UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = f"{titulo_banner} - Sistema Académico UNGE"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="90" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Control Disciplinario y Académico</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: {color_principal};">
                    <h2 style="color: #ffffff; margin: 0; font-size: 18px; font-weight: bold; letter-spacing: 1px; text-transform: uppercase;">{titulo_banner}</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 20px;">{saludo}, <strong>{nombre_alumno}</strong>:</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        {mensaje_estado}
                    </p>

                    <div style="background-color: {color_fondo_caja}; border-radius: 10px; padding: 25px; border: 2px dashed {color_principal}; text-align: center; margin-bottom: 30px;">
                        <p style="margin: 0; font-size: 15px; color: #2c3e50;">{detalle}</p>
                    </div>

                    <p style="font-size: 14px; color: #455a64; margin-bottom: 25px; text-align: center;">
                        Si tiene alguna duda sobre este cambio, por favor contacte con la administración de la facultad.
                    </p>

                    <div style="text-align: center; margin-bottom: 30px;">
                        <a href="{url_btn}" style="background-color: #1a237e; color: #ffffff; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 15px; display: inline-block;">{btn_texto}</a>
                    </div>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Vicerrectorado de Asuntos Académicos - UNGE</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        <span style="font-size: 10px;">Este es un mensaje institucional generado automáticamente por el sistema de gestión.</span>
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        if CORREO_MATRICULAS_PORT == 587:
            server.starttls()
            server.login(CORREO_MATRICULAS_USER, CORREO_MATRICULAS_PASS)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar correo disciplinario: {e}")
        return False

# RUTA PARA CAMBIO DE ESTADO

@app.route('/directivo/cambiar-estado-estudiante/<int:id>', methods=['POST'])
@login_required
def cambiar_estado_estudiante(id):
    # 1. Verificación de seguridad
    if current_user.rol not in ['decano', 'secretaria', 'admin', 'directivo']:
        flash("No tiene permisos para realizar esta acción.", "danger")
        return redirect(url_for('index'))

    estudiante = Usuario.query.get_or_404(id)
    accion = request.form.get('accion') # 'activar' o 'desactivar'
    # Usamos 0 por defecto si no viene el campo meses (en el caso de desactivar)
    meses = int(request.form.get('meses', 0))
    motivo = request.form.get('motivo')

    try:
        if accion == 'activar':
            estudiante.activo = True
            estudiante.motivo_suspension = None
            
            # Lógica de expiración
            if meses > 0:
                estudiante.fecha_expiracion = datetime.utcnow() + timedelta(days=30 * meses)
            else:
                estudiante.fecha_expiracion = None
            
            mensaje = f"Cuenta de {estudiante.nombre} reactivada correctamente."
            categoria = "success"
        
        else: # Acción desactivar
            estudiante.activo = False
            estudiante.fecha_expiracion = None
            estudiante.motivo_suspension = motivo
            mensaje = f"Cuenta de {estudiante.nombre} suspendida."
            categoria = "warning"

        # Guardar en DB
        db.session.commit()

        # Enviar correo (Mailhog capturará esto)
        dominio_app = request.host_url.rstrip('/')
        enviar_correo_disciplinario(
            estudiante=estudiante,
            accion=accion,
            motivo=motivo,
            fecha_expiracion=estudiante.fecha_expiracion,
            dominio=dominio_app
        )

        flash(mensaje, categoria)

    except Exception as e:
        db.session.rollback()
        flash(f"Error al actualizar estado: {str(e)}", "danger")

    return redirect(request.referrer or url_for('panel_control_disciplinario'))



# PANEL PARA MANEJO DE CUENTAS ESTUDIANTILES
@app.route('/directivo/control-disciplinario')
@login_required
def panel_control_disciplinario():
    # 1. Base de la consulta: Solo estudiantes
    query = Usuario.query.filter(Usuario.rol == 'estudiante')

    # 2. Obtener lista de carreras únicas directamente de la DB para el filtro
    # Esto busca todas las carreras diferentes que existen en la tabla usuarios
    carreras_db = db.session.query(Usuario.carrera).filter(
        Usuario.rol == 'estudiante', 
        Usuario.carrera != None
    ).distinct().all()
    
    # Limpiamos la lista (viene como tuplas: [('Medicina',), ('Derecho',)])
    lista_titulaciones = [c[0] for c in carreras_db if c[0]]

    # 3. Aplicar Filtros Dinámicos
    q = request.args.get('q')
    if q:
        # Ahora busca por nombre O apellidos O DIP
        query = query.filter(
            (Usuario.nombre.ilike(f'%{q}%')) | 
            (Usuario.apellidos.ilike(f'%{q}%')) | 
            (Usuario.dip.ilike(f'%{q}%'))
        )

    titulacion = request.args.get('carrera')
    if titulacion:
        query = query.filter(Usuario.carrera == titulacion)

    curso = request.args.get('curso')
    if curso:
        query = query.filter(Usuario.curso == curso)

    estado = request.args.get('estado')
    if estado == 'activo':
        query = query.filter(Usuario.activo == True)
    elif estado == 'inactivo':
        query = query.filter(Usuario.activo == False)

    # 4. Ejecutar consulta
    estudiantes = query.order_by(Usuario.nombre.asc()).all()
    
    return render_template('control_estudiantes.html', 
                           estudiantes=estudiantes, 
                           titulaciones=lista_titulaciones,
                           now=datetime.utcnow())



# ==========================================================
# SECCION DEL ADMINISTRADOR VERSION PANEL
# ==========================================================
# GENERAR UN EXEL PARA RESPALDO DE INFORMACIOn
@app.route('/admin/exportar-usuarios')
def exportar_usuarios_excel():
    try:
        # 1. Consultar todos los usuarios
        usuarios = Usuario.query.all()
        
        if not usuarios:
            flash("No hay usuarios registrados.", "warning")
            return redirect(url_for('panel_admin'))

        # 2. Preparar los datos usando TUS relaciones definidas
        data = []
        for i, u in enumerate(usuarios, 1):
            info_profesional = "N/A"
            seccion_carrera = u.carrera or "N/A"
            
            # Usamos los nombres exactos de tus db.relationship
            if u.rol == 'estudiante' and u.estudiante:
                info_profesional = f"Matrícula: {u.estudiante.matricula}"
                seccion_carrera = u.estudiante.carrera or u.carrera
            
            elif u.rol == 'profesor' and u.profesor:
                info_profesional = f"Título: {u.profesor.titulo_academico}"
                seccion_carrera = u.profesor.departamento
            
            elif u.rol == 'directivo' and u.directivo:
                info_profesional = f"Cargo: {u.directivo.cargo}"
                seccion_carrera = u.directivo.seccion
            
            elif u.rol == 'admin' and u.administrador:
                info_profesional = f"Permisos: {u.administrador.permisos_especiales or 'Total'}"
            
            elif u.rol == 'secretaria' and u.secretaria:
                info_profesional = "Personal de Secretaría"

            # Construcción de la fila con tus campos de base de datos
            data.append({
                'Nº': i,
                'Nombre Completo': f"{u.nombre} {u.apellidos}",
                'DIP': u.dip,
                'Rol': u.rol.upper() if u.rol else 'ESTUDIANTE',
                'Correo Institucional': u.correo_institucional,
                'Teléfono': u.telefono or 'N/A',
                'Sexo': u.sexo or 'N/A',
                'Carrera/Depto': seccion_carrera,
                'Información Extra': info_profesional,
                'Origen': f"{u.natural_de or ''}, {u.distrito_provincia or ''}".strip(', '),
                'Residencia': u.residencia or 'N/A',
                'Debates Part.': u.debate,
                'Estado Cuenta': 'ACTIVO' if u.activo else 'SUSPENDIDO',
                'Fecha de Alta': u.fecha_creacion.strftime('%d/%m/%Y') if u.fecha_creacion else 'N/A'
            })
        
        # 3. Creación del Excel con Estilo UNGE
        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja Principal
            df.to_excel(writer, index=False, sheet_name='BASE DATOS GENERAL')
            
            # Hojas por Categorías
            for rol in ['estudiante', 'profesor', 'directivo', 'admin']:
                df_rol = df[df['Rol'] == rol.upper()].copy()
                if not df_rol.empty:
                    df_rol.loc[:, 'Nº'] = range(1, len(df_rol) + 1)
                    df_rol.to_excel(writer, index=False, sheet_name=f"{rol.capitalize()}s")

            # Estilización
            header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border_side = Side(style='thin', color='000000')
            cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            for sheet in writer.sheets.values():
                for cell in sheet[1]: # Encabezados
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                for row in sheet.iter_rows(): # Bordes para todas las celdas
                    for cell in row:
                        cell.border = cell_border

                # Auto-ajuste de columnas
                for col in sheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    sheet.column_dimensions[col[0].column_letter].width = max_length + 3

        output.seek(0)
        return send_file(output, 
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, 
                         download_name=f"SISTEMA_UNGE_FCS_{datetime.now().strftime('%Y%m%d')}.xlsx")

    except Exception as e:
        print(f"Error: {e}")
        flash(f"Error al generar reporte: {str(e)}", "danger")
        return redirect(url_for('panel_admin'))


# CONSOLA PARA GESTIONAR DB Y STORAGE

@app.route('/admin/consola-recursos')
@login_required # Asegúrate de tener este decorador de seguridad (Mas tarde admin_requiere)
@requiere_rol(['administrador', 'admin'])
def consola_maestra():
    # --- 1. DETECCIÓN DINÁMICA DE TABLAS ---
    # Esto obtiene todas las tablas registradas en tu base de datos automáticamente
    inst = inspect(db.engine)
    nombres_tablas = inst.get_table_names()
    
    tabla_actual = request.args.get('tabla', nombres_tablas[0] if nombres_tablas else None)
    
    columnas = []
    registros = []
    
    if tabla_actual:
        # Buscamos el modelo que coincide con el nombre de la tabla
        model = next((m for m in db.Model.__subclasses__() if m.__tablename__ == tabla_actual), None)
        if model:
            columnas = [c.name for c in model.__table__.columns]
            registros = model.query.all()

    # --- 2. EXPLORADOR DE ARCHIVOS POR CARPETAS ---
    base_path = os.path.join('static', 'uploads')
    # Carpeta en la que estamos actualmente (subdirectorio)
    sub_path = request.args.get('folder', '')
    ruta_completa = os.path.join(base_path, sub_path)
    
    # Seguridad: evitar que el admin suba niveles fuera de uploads
    if not os.path.abspath(ruta_completa).startswith(os.path.abspath(base_path)):
        ruta_completa = base_path

    contenido = []
    if os.path.exists(ruta_completa):
        for item in os.listdir(ruta_completa):
            item_path = os.path.join(ruta_completa, item)
            es_carpeta = os.path.isdir(item_path)
            contenido.append({
                'nombre': item,
                'es_dir': es_carpeta,
                'ruta_relativa': os.path.join(sub_path, item),
                'size': os.path.getsize(item_path) // 1024 if not es_carpeta else 0
            })

    return render_template('consola_admin.html', 
                           tablas=nombres_tablas, 
                           tabla_actual=tabla_actual,
                           columnas=columnas, 
                           registros=registros,
                           archivos=contenido,
                           folder_actual=sub_path)

# --- 3. ACCIONES (API PARA LOS BOTONES) ---
@app.route('/admin/api/delete-record', methods=['POST'])
def delete_record():
    data = request.json
    # Lógica para borrar de la DB: db.session.delete(...)
    return jsonify({"status": "success", "message": "Registro eliminado"})

@app.route('/admin/api/delete-file', methods=['POST'])
def delete_file():
    data = request.json
    # Lógica para borrar archivo físico: os.remove(...)
    return jsonify({"status": "success", "message": "Archivo eliminado"})


# Devolver JSON para modificarlos (opcional)
@app.route('/admin/api/get-record/<tabla>/<int:id>')
def get_record(tabla, id):
    # Buscar el modelo dinámicamente
    model = next((m for m in db.Model.__subclasses__() if m.__tablename__ == tabla), None)
    if not model:
        return jsonify({"error": "Tabla no encontrada"}), 404
    
    registro = model.query.get_or_404(id)
    
    # Convertir el registro a un diccionario (evitando datos sensibles si es necesario)
    columnas = [c.name for c in model.__table__.columns]
    datos = {col: getattr(registro, col) for col in columnas if col != 'password_hash'}
    
    return jsonify(datos)

@app.route('/admin/api/update-record/<tabla>/<int:id>', methods=['POST'])
def update_record(tabla, id):
    data = request.json
    model = next((m for m in db.Model.__subclasses__() if m.__tablename__ == tabla), None)
    registro = model.query.get(id)
    
    for key, value in data.items():
        if hasattr(registro, key):
            setattr(registro, key, value)
    
    db.session.commit()
    return jsonify({"status": "success"})

#--------------------------------------------------\\
# Cambiar contrasena
@app.route('/admin/api/soporte/generar-recuperacion', methods=['POST'])
@login_required
def generar_recuperacion_soporte():
    data = request.json
    email_inst = data.get('email')
    
    # Buscamos al usuario por correo_institucional (asumiendo que tu modelo se llama Usuario)
    usuario = Usuario.query.filter_by(correo_institucional=email_inst).first()
    
    if not usuario:
        return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

        # Verificamos si tiene el correo personal (llamado 'correo')
    if not usuario.correo or "@" not in usuario.correo:
        return jsonify({
            "status": "error", 
            "message": "El usuario no tiene un correo personal válido registrado para recibir el código."
        }), 400

    # Lógica: Generar un código temporal o resetear contraseña
    nuevo_codigo = secrets.token_hex(4).upper() # Ejemplo: A1B2C3D4
    
    # Aquí puedes:
    # 1. Guardar el código en la base de datos
    # 2. Enviar un correo real al usuario
    # 3. O simplemente devolverlo para que el Admin se lo de por teléfono
    
    return jsonify({
        "status": "success", 
        "message": f"Código generado: {nuevo_codigo}. Se ha enviado a {usuario.correo}."
    })

@app.route('/admin/api/delete-record/<tabla>/<int:id>', methods=['POST'])
@login_required
def delete_record_full(tabla, id):
    # Buscar el modelo dinámicamente
    model = next((m for m in db.Model.__subclasses__() if m.__tablename__ == tabla), None)
    if not model:
        return jsonify({"status": "error", "message": "Tabla no encontrada"}), 404
    
    registro = model.query.get(id)
    if registro:
        db.session.delete(registro)
        db.session.commit()
        return jsonify({"status": "success", "message": "Registro eliminado permanentemente"})
    return jsonify({"status": "error", "message": "No se encontró el registro"}), 404
# --------------------------------------------------
# Recuperar cuentas con correos automaticos
# --------------------------------------------------
# Ruta para generar recuperación desde la consola
# --------------------------------------------------
@app.route('/amin/api/soporte/generar-recuperacion', methods=['POST'])
def generar_recuperacion():
    print(">>> PETICIÓN RECIBIDA") 
    
    try:
        data = request.get_json()
        email = data.get('email')
        print(f">>> Buscando en DB el correo personal: {email}")

        # CAMBIO AQUÍ: Usamos 'correo' en lugar de 'correo_institucional'
        usuario = Usuario.query.filter_by(correo=email).first()

        if not usuario:
            print(">>> Usuario no encontrado con ese correo personal")
            return jsonify({"status": "error", "message": "Correo no registrado"}), 404

        codigo = secrets.token_hex(3).upper() 
        
        # Guardar en DB
        usuario.recovery_code = codigo
        usuario.recovery_expire = datetime.utcnow() + timedelta(minutes=30)
        db.session.commit()

        dominio = request.host_url.rstrip('/')
        # Pasamos el usuario a la función de envío
        exito = enviar_recuperacion_consola(usuario, codigo, dominio)

        if exito:
            return jsonify({"status": "success", "message": "Código enviado a MailHog"})
        else:
            return jsonify({"status": "error", "message": "Fallo en conexión SMTP"}), 500

    except Exception as e:
        print(f">>> ERROR: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


#------------------------------------------------------------------
# Enviar correo de recuperacion
def enviar_recuperacion_consola(usuario, codigo_temporal, dominio):
    """
    Envía un correo de soporte desde la consola administrativa para recuperar acceso.
    Adaptado a la estructura institucional de la UNGE y configuración MailHog.
    """
    email_destino = usuario.correo
    nombre_usuario = f"{usuario.nombre} {usuario.apellidos}"
    
    msg = MIMEMultipart('alternative')
    msg['From'] = f"Soporte IT UNGE <{CORREO_MATRICULAS_USER}>"
    msg['To'] = email_destino
    msg['Subject'] = "Restablecimiento de Acceso - Consola Administrativa"

    # Enlace a la página de cambio de contraseña o login
    url_soporte = f"{dominio}/reset-password?email={usuario.correo}&codigo={codigo_temporal}"

    html_content = f"""
    <html>
    <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7f9;">
        <table align="center" border="0" cellpadding="0" cellspacing="0" width="600" style="border-collapse: collapse; background-color: #ffffff; margin-top: 30px; border-radius: 15px; border: 1px solid #e1e8ed; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">
            
            <tr>
                <td align="center" style="padding: 40px 0 20px 0; background-color: #ffffff;">
                    <img src="{dominio}/static/img/logo_unge.jpeg" alt="Logo UNGE" width="100" style="display: block; margin-bottom: 15px;">
                    <h1 style="margin: 0; font-size: 14px; color: #1a237e; letter-spacing: 1px; text-transform: uppercase;">Universidad Nacional de Guinea Ecuatorial</h1>
                    <p style="margin: 5px 0 0 0; font-size: 16px; color: #ff6f00; font-weight: bold;">Centro de Soporte Técnico</p>
                </td>
            </tr>

            <tr>
                <td style="padding: 20px 40px; text-align: center; background-color: #1a237e;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 300; letter-spacing: 1px;">RECUPERACIÓN DE CUENTA</h2>
                </td>
            </tr>

            <tr>
                <td style="padding: 40px 40px 20px 40px;">
                    <p style="font-size: 18px; color: #2c3e50; margin-bottom: 25px;">Estimado(a) <strong>{nombre_usuario}</strong>,</p>
                    
                    <p style="font-size: 15px; color: #546e7a; line-height: 1.6; margin-bottom: 25px;">
                        Un administrador de la facultad ha generado un <strong>Código de Acceso Temporal</strong> para ayudarle a recuperar el acceso a su cuenta institucional de la FCS. 
                    </p>

                    <div style="background-color: #fff9e6; border-radius: 10px; padding: 25px; border: 2px dashed #ffc107; text-align: center; margin-bottom: 30px;">
                        <p style="margin: 0 0 10px 0; font-size: 13px; color: #856404; font-weight: bold; text-transform: uppercase;">Su código temporal de seguridad es:</p>
                        <span style="font-size: 32px; font-weight: bold; color: #1a237e; letter-spacing: 4px;">{codigo_temporal}</span>
                    </div>

                    <p style="font-size: 15px; color: #455a64; margin-bottom: 25px; text-align: center;">
                        Utilice este código en el portal de soporte para establecer una nueva contraseña de acceso:
                    </p>

                    <div style="text-align: center; margin-bottom: 30px;">
                        <a href="{url_soporte}" style="background-color: #ff6f00; color: #ffffff; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; font-size: 16px; display: inline-block;">RECUPERAR MI ACCESO</a>
                    </div>

                    <p style="font-size: 13px; color: #78909c; text-align: center; font-style: italic;">
                        Este código expirará en 24 horas por razones de seguridad. Si usted no ha solicitado esta ayuda, ignore este mensaje.
                    </p>
                </td>
            </tr>

            <tr>
                <td style="background-color: #eceff1; padding: 30px; text-align: center; border-top: 1px solid #cfd8dc;">
                    <p style="margin: 0; color: #78909c; font-size: 12px; line-height: 1.5;">
                        <strong>Soporte IT - Facultad de Ciencias de la Salud</strong><br>
                        Campus de Bata, Guinea Ecuatorial<br>
                        Este es un mensaje automático del Sistema de Gestión UNGE.
                    </p>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_content, 'html'))

    try:
        # Conexión a MailHog
        server = smtplib.SMTP(CORREO_MATRICULAS_SERVER, CORREO_MATRICULAS_PORT)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Error al enviar recuperación de cuenta: {e}")
        return False


# _------------------------------------------
# Finalizar cambio
@app.route('/reset-password', methods=['POST'])
def finalizar_restablecimiento():
    email_personal = request.form.get('email')
    codigo_ingresado = request.form.get('codigo')
    nueva_pass = request.form.get('nueva_pass')

    # 1. Búsqueda exacta
    usuario = Usuario.query.filter_by(correo=email_personal).first()

    if not usuario:
        return jsonify({"status": "error", "message": "Identidad no encontrada"}), 404
        
    if usuario.recovery_code != codigo_ingresado:
        return jsonify({"status": "error", "message": "Token inválido"}), 400

    try:
        # 2. GENERAR EL HASH (Asegúrate de importar generate_password_hash)
        nuevo_hash = generate_password_hash(nueva_pass)
        
        # DEBUG: Imprime en consola para verificar que el hash cambia
        print(f"DEBUG: Actualizando pass para {usuario.correo_institucional}")
        print(f"DEBUG: Hash anterior: {usuario.password_hash[:15]}...")
        
        # 3. ASIGNACIÓN DIRECTA
        usuario.password_hash = nuevo_hash
        
        # 4. LIMPIEZA DE SEGURIDAD
        usuario.recovery_code = None
        usuario.recovery_expire = None
        
        # 5. COMMIT FORZADO
        db.session.add(usuario) # Aseguramos que el objeto esté en la sesión
        db.session.commit()
        
        print(f"DEBUG: ¡Éxito! Nueva clave guardada para login institucional.")

        return jsonify({
            "status": "success", 
            "message": f"Acceso actualizado. Usa tu correo: {usuario.correo_institucional}",
            "redirect": "/login" # Ajusta a tu ruta real de login
        })

    except Exception as e:
        db.session.rollback()
        print(f"DEBUG ERROR: {str(e)}")
        return jsonify({"status": "error", "message": "Error interno al guardar"}), 500

# Formulario para crear nueva contrasena.
@app.route('/reset-password', methods=['GET'])
def mostrar_formulario_reset():
    # Capturamos los datos que vienen en el enlace del correo
    email = request.args.get('email', '')
    codigo = request.args.get('codigo', '')
    
    # Renderizamos el HTML pasando estos valores para que aparezcan rellenos
    return render_template('reset_password.html', email=email, codigo=codigo)


# -----------------------------------------------------------------
# CONTROL DE CPU Y BD SEGUN CONSUMO

@app.route('/admin/api/stats')
@login_required
def get_stats():
    # 1. Signos vitales universales (CPU y RAM)
    cpu = psutil.cpu_percent(interval=0.1)
    ram = psutil.virtual_memory().percent
    
    # 2. Base de Datos
    try:
        db.session.execute(text('SELECT 1'))
        db_status = "En línea"
    except Exception:
        db_status = "Desconectada"

    # 3. Almacenamiento Inteligente (Detección de Entorno)
    # Si detectamos que estamos en Render o VPS, usamos la raíz, 
    # si no, usamos la ruta donde se ejecuta el script.
    ruta_a_medir = '/' if os.name != 'nt' else os.getcwd()
    
    try:
        total, used, free = shutil.disk_usage(ruta_a_medir)
        disco_p = round((used / total) * 100, 1)
        # Formatear espacio libre de forma legible
        libre_gb = round(free / (1024**3), 2)
    except:
        disco_p = 0
        libre_gb = 0

    return jsonify({
        "cpu": cpu,
        "ram": ram,
        "db": db_status,
        "disco": disco_p,
        "disco_info": f"{libre_gb} GB libres",
        "entorno": "Producción" if os.name != 'nt' else "Local/Desarrollo"
    })

# ==========================================================
# CONFIGURACIÓN DE FLASK-LOGIN (PROFESIONAL)
# ==========================================================
login_manager = LoginManager()
login_manager.init_app(app)

# Indicar la ruta del login (ajusta 'login_page' al nombre real de tu función)
login_manager.login_view = 'login_page' 
login_manager.login_message = "Por favor, inicia sesión para acceder."
login_manager.login_message_category = "info"

@login_manager.user_loader
def load_user(user_id):
    # Flask-Login usará esta función para cargar al usuario en 'current_user'
    print(f"DEBUG: Cargando usuario ID {user_id}") # Esto saldrá en tu terminal
    return Usuario.query.get(int(user_id))






# ==========================================================
# INICIAR SERVIDOR
# ==========================================================
with app.app_context():
    # db.create_all()
    pass
    

if __name__ == '__main__':
    app.run(host="127.0.0.1", port=5000, debug=True)

